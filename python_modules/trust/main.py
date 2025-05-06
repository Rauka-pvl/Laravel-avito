import os
import re
import time
import random
import logging
import requests
import shutil
from time import sleep
from datetime import datetime
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import mysql.connector
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager

# === Импорт конфигурации из avito ===
import sys
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "avito")))
from config import COMBINED_XML, LOG_DIR, BASE_DIR

# === Пути ===
LOG_DIR = os.path.join(BASE_DIR, "..", ".." ,"storage", "app", "public", "log-trast")
OUTPUT_FILE = os.path.join(LOG_DIR, "..", "trast.xlsx")
BACKUP_FILE = os.path.join(LOG_DIR, "..", "trast_backup.xlsx")
os.makedirs(LOG_DIR, exist_ok=True)

# === Логирование ===
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
log_filename = os.path.join(LOG_DIR, f"trast_{timestamp}.log")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filename, encoding="utf-8"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# === Бэкап Excel ===
def create_backup():
    if os.path.exists(OUTPUT_FILE):
        shutil.copy2(OUTPUT_FILE, BACKUP_FILE)
        logger.info(f"Бэкап создан: {BACKUP_FILE}")

# === Подключение к БД ===
def connect_to_db(retries=3, delay=3):
    for attempt in range(retries):
        try:
            conn = mysql.connector.connect(
                host="127.0.0.1",
                user="uploader",
                password="uploader",
                database="avito"
            )
            if conn.is_connected():
                logger.info("Успешное подключение к базе данных")
                return conn
        except mysql.connector.Error as err:
            logger.warning(f"Попытка {attempt+1}: ошибка подключения к БД: {err}")
        sleep(delay)
    logger.warning("База данных недоступна.")
    return None

def update_config_status(db, name, value):
    if not db:
        return
    try:
        with db.cursor() as cursor:
            cursor.execute("SELECT COUNT(*) FROM config WHERE name = %s", (name,))
            exists = cursor.fetchone()[0]
            if exists:
                cursor.execute("UPDATE config SET value = %s WHERE name = %s", (value, name))
            else:
                cursor.execute("INSERT INTO config (name, value) VALUES (%s, %s)", (name, value))
            db.commit()
    except Exception as e:
        db.rollback()
        logger.error(f"Ошибка при обновлении конфигурации: {e}")

# === Selenium Driver ===
def create_driver():
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-gpu")
    return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# === Получение количества страниц ===
def get_pages_count(url="https://trast-zapchast.ru/shop/"):
    driver = create_driver()
    try:
        driver.get(url)
        time.sleep(3)
        soup = BeautifulSoup(driver.page_source, "html.parser")
        pages = soup.select("div.th-products-view__pagination ul.page-numbers a")
        numbers = [int(a.text.strip()) for a in pages if a.text.strip().isdigit()]
        return max(numbers, default=1)
    finally:
        driver.quit()

# === Получение ссылок на карточки ===
def get_product_links(page_url):
    driver = create_driver()
    try:
        driver.get(page_url)
        time.sleep(3)
        soup = BeautifulSoup(driver.page_source, "html.parser")
        cards = soup.select("li.product:not(.outofstock)")
        return [a["href"] for card in cards if (a := card.find("a", class_="woocommerce-LoopProduct-link"))]
    finally:
        driver.quit()

# === Безопасный GET с ретраями ===
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/123.0.0.0 Safari/537.36"
}

def safe_get(url, retries=3, delay=2):
    for i in range(retries):
        try:
            response = requests.get(url, headers=HEADERS, timeout=10)
            if response.status_code == 200:
                return response
            else:
                logger.warning(f"Статус {response.status_code} от {url}")
        except Exception as e:
            logger.warning(f"Попытка {i+1}: ошибка {e}")
        time.sleep(delay)
    return None

# === Парсинг карточки товара ===
def parse_product_page_single_price(url):
    response = safe_get(url)
    if not response:
        return {}

    soup = BeautifulSoup(response.content, "html.parser")
    data = {
        "manufacturer": None,
        "article": None,
        "description": None,
        "price": None,
        "analogs": None,
    }

    title = soup.select_one("h1.product_title.entry-title")
    if title:
        data["description"] = title.text.strip()

    attrs = soup.select("div.wl-attr--list .wl-attr--item")
    for attr in attrs:
        text = attr.get_text(strip=True)
        val = attr.select_one(".pa-right")
        if not val:
            continue
        value = val.text.strip()
        if "Производитель" in text:
            data["manufacturer"] = value
        elif "Артикул" in text:
            data["article"] = value
        elif "Аналоги" in text:
            data["analogs"] = value

    price = soup.select_one("div.wl-variable--item .wl-variable--price")
    if price:
        data["price"] = {"price": re.sub(r"[^\d]", "", price.text.strip())}
    return data

# === Работа с Excel ===
def create_new_excel(path):
    if os.path.exists(path):
        os.remove(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(["Производитель", "Артикул", "Описание", "Цена", "Аналоги"])
    wb.save(path)

def append_to_excel(path, product_list):
    wb = load_workbook(path)
    ws = wb.active
    for p in product_list:
        ws.append([
            p.get("manufacturer", ""),
            p.get("article", ""),
            p.get("description", ""),
            p.get("price", {}).get("price", ""),
            p.get("analogs", "")
        ])
    wb.save(path)

# === Запуск ===
if __name__ == "__main__":
    start = time.time()
    create_backup()
    create_new_excel(OUTPUT_FILE)
    db = connect_to_db()

    try:
        if db:
            update_config_status(db, 'parser_status', 'in_progress')
            update_config_status(db, 'parser_update_time', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

        links = []
        for page_num in range(1, get_pages_count() + 1):
            page_url = f"https://trast-zapchast.ru/shop/page/{page_num}/"
            links += get_product_links(page_url)

        for link in links:
            try:
                product = parse_product_page_single_price(link)
                if product:
                    append_to_excel(OUTPUT_FILE, [product])
                time.sleep(random.uniform(0.5, 1.5))
            except Exception as e:
                logger.error(f"Ошибка при обработке ссылки {link}: {e}")

        if db:
            update_config_status(db, 'parser_status', 'done')
            update_config_status(db, 'parser_update_time', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    except Exception as e:
        logger.critical(f"Фатальная ошибка: {e}")
        if db:
            update_config_status(db, 'parser_status', 'failed')
            update_config_status(db, 'parser_update_time', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    finally:
        if db:
            db.close()

    logger.info(f"Парсинг завершён за {round(time.time() - start, 2)} сек.")
