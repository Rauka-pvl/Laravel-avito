from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import os
import logging
import requests
import re
import mysql.connector
from mysql.connector import Error
from openpyxl import Workbook, load_workbook
import time
import random
from time import sleep
from concurrent.futures import ThreadPoolExecutor, as_completed

# === Константы ===
LOGS_PATH = "/home/admin/web/233204.fornex.cloud/public_html/storage/logs/update/"
OUTPUT_FILE = "/home/admin/web/233204.fornex.cloud/public_html/public/products.xlsx"
DB_CONFIG = {
    "host": "127.0.0.1",
    "user": "uploader",
    "password": "uploader",
    "database": "avito"
}

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36"
]

PROXY_LIST = [
    "212.113.123.246:42681", "194.87.201.123:24645", "proxy.vpnnaruzhu.com:60000:guest2:o7PwR99l"
]

# === Логирование ===
os.makedirs(LOGS_PATH, exist_ok=True)
log_file = os.path.join(LOGS_PATH, f"parser-2_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(log_file, encoding="utf-8"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# === Вспомогательные функции ===
def get_random_proxy():
    proxy = random.choice(PROXY_LIST)
    if proxy.count(":") == 1:
        return {"http": f"http://{proxy}", "https": f"http://{proxy}"}
    if proxy.count(":") == 3:
        host, port, user, password = proxy.split(":")
        auth = f"{user}:{password}@{host}:{port}"
        return {"http": f"http://{auth}", "https": f"http://{auth}"}
    raise ValueError(f"Некорректный формат прокси: {proxy}")

def fetch_with_proxy(url, max_attempts=5, timeout=10):
    for _ in range(max_attempts):
        try:
            proxy = get_random_proxy()
            headers = {"User-Agent": random.choice(USER_AGENTS)}
            response = requests.get(url, headers=headers, proxies=proxy, timeout=timeout)
            response.raise_for_status()
            return response
        except Exception:
            sleep(1)
    return None

def fetch_with_fallback_proxy(url, timeout=10):
    try:
        headers = {"User-Agent": random.choice(USER_AGENTS)}
        response = requests.get(url, headers=headers, timeout=timeout)
        response.raise_for_status()
        return response
    except Exception:
        return fetch_with_proxy(url, timeout=timeout)

def connect_to_db():
    return mysql.connector.connect(**DB_CONFIG)

def update_config_status(db, name, value):
    if db is None:
        logger.warning(f"Пропуск обновления '{name}' — нет подключения к БД")
        return
    try:
        with db.cursor() as cursor:
            cursor.execute("SELECT COUNT(*) FROM config WHERE name=%s", (name,))
            if cursor.fetchone()[0]:
                cursor.execute("UPDATE config SET value=%s WHERE name=%s", (value, name))
            else:
                cursor.execute("INSERT INTO config (name, value) VALUES (%s, %s)", (name, value))
        db.commit()
    except Exception as e:
        logger.error(f"DB error: {e}")
        try:
            db.rollback()
        except:
            pass

def create_new_excel(path):
    if os.path.exists(path): os.remove(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(["Производитель", "Артикул", "Описание", "Цена", "Аналоги"])
    wb.save(path)

def append_products_to_excel(path, products):
    wb = load_workbook(path)
    ws = wb.active
    for p in products:
        ws.append([
            p.get("manufacturer", ""),
            p.get("article", ""),
            p.get("description", ""),
            p.get("price", {}).get("price", ""),
            p.get("analogs", "")
        ])
    wb.save(path)

def clean_price(text):
    return re.sub(r"[^\d]", "", text)

def parse_product_page_single_price(url):
    result = {"manufacturer": None, "article": None, "description": None, "price": None, "analogs": None}
    r = fetch_with_fallback_proxy(url)
    if not r: return result
    soup = BeautifulSoup(r.text, "html.parser")

    title = soup.select_one("h1.product_title.entry-title")
    if title: result["description"] = title.text.strip()

    for attr in soup.select("div.wl-attr--list .wl-attr--item"):
        label = attr.get_text(strip=True)
        value = attr.select_one(".pa-right")
        if not value: continue
        value = value.text.strip()
        if "Производитель" in label:
            result["manufacturer"] = value
        elif "Артикул" in label:
            result["article"] = value
        elif "Аналоги" in label:
            result["analogs"] = value

    price = soup.select_one("div.wl-variable--item .wl-variable--price")
    if price:
        result["price"] = {"price": clean_price(price.text)}

    return result

def get_pages_count(base_url="https://trast-zapchast.ru/shop/"):
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    try:
        driver.get(base_url)
        time.sleep(3)
        soup = BeautifulSoup(driver.page_source, "html.parser")
        pagination = soup.select_one("div.th-products-view__pagination ul.page-numbers")
        return max((int(a.text) for a in pagination.find_all("a") if a.text.isdigit()), default=1)
    finally:
        driver.quit()

def get_product_links(url, driver):
    driver.get(url)
    time.sleep(3)
    soup = BeautifulSoup(driver.page_source, "html.parser")
    products = []
    for card in soup.select("li.product"):
        if "outofstock" in card.get("class", []): continue
        link = card.select_one("a.woocommerce-LoopProduct-link")
        if link: products.append(link.get("href"))
    return products

def main():
    db = None
    os.environ["WDM_LOCAL"] = "1"
    start = time.time()
    create_new_excel(OUTPUT_FILE)
    links = []

    try:
        try:
            db = connect_to_db()
            update_config_status(db, 'parser_status', 'in_progress')
            update_config_status(db, 'parser_update_time', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        except Exception as db_err:
            logger.error(f"Не удалось подключиться к БД: {db_err}")
            db = None

        # Кэшированный ChromeDriver
        options = Options()
        options.add_argument("--headless")
        options.add_argument("--disable-gpu")
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
        try:
            for page_num in range(1, get_pages_count() + 1):
                links.extend(get_product_links(f"https://trast-zapchast.ru/shop/page/{page_num}/", driver))
        finally:
            driver.quit()

        # Параллельная обработка товаров
        with ThreadPoolExecutor(max_workers=5) as executor:
            logger.info(f"Начинаем парсинг {len(links)} ссылок в потоках...")
            futures = [executor.submit(parse_product_page_single_price, url) for url in links]
            products = []
            for f in as_completed(futures, timeout=180):
                try:
                    result = f.result(timeout=15)
                    if result and result.get("description"):
                        products.append(result)
                    else:
                        logger.warning("Пустой результат, пропущен.")
                except Exception as e:
                    logger.warning(f"Ошибка при обработке карточки: {e}")

        logger.info(f"Успешно обработано {len(products)} товаров. Записываем в Excel...")
        append_products_to_excel(OUTPUT_FILE, products)

        if db:
            update_config_status(db, 'parser_status', 'done')
    except Exception as e:
        logger.exception("Ошибка выполнения")
        if db:
            update_config_status(db, 'parser_status', 'failed')
    finally:
        if db:
            update_config_status(db, 'parser_update_time', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    logger.info(f"Завершено за {round(time.time() - start, 2)} сек")

if __name__ == "__main__":
    main()
