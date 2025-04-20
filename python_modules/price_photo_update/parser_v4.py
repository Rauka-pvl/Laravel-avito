from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import logging
import requests
import re
import mysql.connector
from mysql.connector import Error
import time
import random
import random
import requests
from time import sleep

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s'
)
logger = logging.getLogger(__name__)

USER_AGENTS = [
    # Chrome –Ω–∞ Windows
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    
    # Firefox –Ω–∞ Windows
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:124.0) Gecko/20100101 Firefox/124.0",

    # Chrome –Ω–∞ macOS
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.6312.107 Safari/537.36",

    # Safari –Ω–∞ macOS
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_2_1) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.3 Safari/605.1.15",

    # Chrome –Ω–∞ Android
    "Mozilla/5.0 (Linux; Android 13; Pixel 6 Pro) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Mobile Safari/537.36",

    # Samsung Internet –Ω–∞ Android
    "Mozilla/5.0 (Linux; Android 12; SAMSUNG SM-G998B) AppleWebKit/537.36 (KHTML, like Gecko) SamsungBrowser/20.0 Chrome/96.0.4664.45 Mobile Safari/537.36",

    # Firefox –Ω–∞ Android
    "Mozilla/5.0 (Android 13; Mobile; rv:110.0) Gecko/110.0 Firefox/110.0",

    # iPhone Safari
    "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Mobile/15E148 Safari/604.1",

    # Googlebot
    "Mozilla/5.0 (compatible; Googlebot/2.1; +http://www.google.com/bot.html)",

    # Bingbot
    "Mozilla/5.0 (compatible; bingbot/2.0; +http://www.bing.com/bingbot.htm)"
]
PROXY_LIST = [
    "212.113.123.246:42681",
    "194.87.201.123:24645",
    "92.53.127.107:27807",
    "79.137.133.95:40764",
    "176.124.217.180:12048",
    "92.63.176.36:36162",
    "147.45.254.230:45237",
    "193.164.150.44:57274",
    "89.23.116.215:28305",
    "185.247.185.200:35947",
    "92.255.109.228:25329",
    "176.124.215.172:17789",
    "176.124.216.190:47804",
    "147.45.70.59:47706",
    "proxy.vpnnaruzhu.com:60000:guest2:o7PwR99l",
    "46.232.124.235:50100:bot0TN93:wLiPe9hNN8",
    "149.126.231.118:50100:bot0TN93:wLiPe9hNN8",
    "74.124.221.75:50100:bot0TN93:wLiPe9hNN8",
    "209.200.239.109:50100:bot0TN93:wLiPe9hNN8",
    "45.135.38.65:50100:bot0TN93:wLiPe9hNN8"
]

def get_random_proxy():
    proxy = random.choice(PROXY_LIST)
    if proxy.count(":") == 1:
        return {
            "http": f"http://{proxy}",
            "https": f"http://{proxy}"
        }
    elif proxy.count(":") == 3:
        host, port, user, password = proxy.split(":")
        return {
            "http": f"http://{user}:{password}@{host}:{port}",
            "https": f"http://{user}:{password}@{host}:{port}"
        }
    else:
        raise ValueError(f"–ù–µ–∫–æ—Ä—Ä–µ–∫—Ç–Ω—ã–π —Ñ–æ—Ä–º–∞—Ç –ø—Ä–æ–∫—Å–∏: {proxy}")

def fetch_with_proxy(url: str, max_attempts: int = 10, timeout: int = 10) -> requests.Response | None:
    for attempt in range(max_attempts):
        try:
            proxy = get_random_proxy()
            print(f"[{attempt+1}] –ü—Ä–æ–±—É–µ–º –ø—Ä–æ–∫—Å–∏: {proxy['http']}")
            response = requests.get(url, proxies=proxy, timeout=timeout)
            response.raise_for_status()
            print("‚úÖ –£—Å–ø–µ—à–Ω–æ–µ –ø–æ–¥–∫–ª—é—á–µ–Ω–∏–µ")
            return response
        except Exception as e:
            print(f"‚ùå –û—à–∏–±–∫–∞ —Å –ø—Ä–æ–∫—Å–∏: {e}")
            sleep(1)  # –ú–æ–∂–Ω–æ —É–±—Ä–∞—Ç—å –∏–ª–∏ —É–≤–µ–ª–∏—á–∏—Ç—å –ø–∞—É–∑—É
    print("üö´ –ù–µ —É–¥–∞–ª–æ—Å—å –ø–æ–¥–∫–ª—é—á–∏—Ç—å—Å—è –ø–æ—Å–ª–µ –≤—Å–µ—Ö –ø–æ–ø—ã—Ç–æ–∫")
    return None

def fetch_with_fallback_proxy(url: str, timeout: int = 10):
    try:
        logger.info(f"–ü—Ä–æ–±—É–µ–º –∑–∞–≥—Ä—É–∑–∏—Ç—å —Å—Ç—Ä–∞–Ω–∏—Ü—É –±–µ–∑ –ø—Ä–æ–∫—Å–∏: {url}")
        response = requests.get(url, timeout=timeout, headers={"User-Agent": random.choice(USER_AGENTS)})
        response.raise_for_status()
        return response
    except Exception as e:
        logger.warning(f"–û—à–∏–±–∫–∞ –±–µ–∑ –ø—Ä–æ–∫—Å–∏: {e}. –ü—Ä–æ–±—É–µ–º —Å –ø—Ä–æ–∫—Å–∏...")
        return fetch_with_proxy(url, timeout)

def get_random_headers():
    return {"User-Agent": random.choice(USER_AGENTS)}


def get_pages_count(url: str = "https://trast-zapchast.ru/shop/") -> int:
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    from webdriver_manager.chrome import ChromeDriverManager
    from bs4 import BeautifulSoup
    import time

    logger.info(f"–ó–∞–ø—É—Å–∫ —Ñ—É–Ω–∫—Ü–∏–∏ get_pages_count –¥–ª—è URL: {url}")

    options = Options()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    try:
        logger.info("–ó–∞–≥—Ä—É–∂–∞–µ–º —Å—Ç—Ä–∞–Ω–∏—Ü—É –≤ –±—Ä–∞—É–∑–µ—Ä–µ")
        driver.get(url)
        time.sleep(3)

        logger.info("–ü–æ–ª—É—á–∞–µ–º HTML —Å—Ç—Ä–∞–Ω–∏—Ü—ã")
        soup = BeautifulSoup(driver.page_source, "html.parser")
        pagination = soup.select_one("div.th-products-view__pagination ul.page-numbers")
        last_page = 1

        if pagination:
            logger.info("–û–±–Ω–∞—Ä—É–∂–µ–Ω –±–ª–æ–∫ –ø–∞–≥–∏–Ω–∞—Ü–∏–∏, –Ω–∞—á–∏–Ω–∞–µ–º –æ–±—Ä–∞–±–æ—Ç–∫—É —Å—Å—ã–ª–æ–∫")
            page_links = pagination.find_all("a", class_="page-numbers")
            for link in page_links:
                try:
                    page_num = int(link.text.strip())
                    logger.info(f"–û–±–Ω–∞—Ä—É–∂–µ–Ω–∞ —Å—Ç—Ä–∞–Ω–∏—Ü–∞: {page_num}")
                    last_page = max(last_page, page_num)
                except ValueError:
                    logger.info(f"–ü—Ä–æ–ø—É—â–µ–Ω —ç–ª–µ–º–µ–Ω—Ç –ø–∞–≥–∏–Ω–∞—Ü–∏–∏: {link.text.strip()}")
        else:
            logger.warning("–ë–ª–æ–∫ –ø–∞–≥–∏–Ω–∞—Ü–∏–∏ –Ω–µ –Ω–∞–π–¥–µ–Ω")

        logger.info(f"–û–±—â–µ–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ —Å—Ç—Ä–∞–Ω–∏—Ü: {last_page}")
        return last_page

    finally:
        driver.quit()
        logger.info("–î—Ä–∞–π–≤–µ—Ä –∑–∞–∫—Ä—ã—Ç")

def get_product_links(page_url: str) -> list[str]:
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    from webdriver_manager.chrome import ChromeDriverManager
    from bs4 import BeautifulSoup
    import time
    import logging

    logger = logging.getLogger(__name__)
    logger.info(f"–ó–∞–ø—É—Å–∫ —Ñ—É–Ω–∫—Ü–∏–∏ get_product_links –¥–ª—è —Å—Ç—Ä–∞–Ω–∏—Ü—ã: {page_url}")

    options = Options()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    try:
        logger.info("–ó–∞–≥—Ä—É–∂–∞–µ–º —Å—Ç—Ä–∞–Ω–∏—Ü—É –≤ –±—Ä–∞—É–∑–µ—Ä–µ")
        driver.get(page_url)
        time.sleep(3)

        logger.info("–ü–æ–ª—É—á–∞–µ–º HTML —Å—Ç—Ä–∞–Ω–∏—Ü—ã")
        soup = BeautifulSoup(driver.page_source, "html.parser")
        product_links = []

        logger.info("–ò—â–µ–º –≤—Å–µ –∫–∞—Ä—Ç–æ—á–∫–∏ —Ç–æ–≤–∞—Ä–æ–≤")
        product_cards = soup.select("li.product")

        logger.info(f"–û–±–Ω–∞—Ä—É–∂–µ–Ω–æ –∫–∞—Ä—Ç–æ—á–µ–∫ —Ç–æ–≤–∞—Ä–æ–≤: {len(product_cards)}")

        for i, product in enumerate(product_cards, 1):
            class_list = product.get("class", [])

            if 'outofstock' in class_list:
                logger.info(f"[{i}] –ü—Ä–æ–ø—É—â–µ–Ω ‚Äî —Ç–æ–≤–∞—Ä –Ω–µ –≤ –Ω–∞–ª–∏—á–∏–∏ (class —Å–æ–¥–µ—Ä–∂–∏—Ç 'outofstock')")
                continue

            a_tag = product.find("a", class_="woocommerce-LoopProduct-link")
            href = a_tag.get("href") if a_tag else None

            if href:
                product_links.append(href)
                logger.info(f"[{i}] –î–æ–±–∞–≤–ª–µ–Ω–∞ —Å—Å—ã–ª–∫–∞: {href}")
            else:
                logger.info(f"[{i}] –°—Å—ã–ª–∫–∞ –Ω–µ –Ω–∞–π–¥–µ–Ω–∞")

        logger.info(f"–í—Å–µ–≥–æ –¥–æ–±–∞–≤–ª–µ–Ω–æ —Å—Å—ã–ª–æ–∫: {len(product_links)}")
        return product_links

    finally:
        driver.quit()
        logger.info("–î—Ä–∞–π–≤–µ—Ä –∑–∞–∫—Ä—ã—Ç")

def clean_price(price_str: str) -> str:
    return re.sub(r"[^\d]", "", price_str)

def parse_product_page_single_price(url: str) -> dict:
    logger.info(f"–ó–∞–≥—Ä—É–∂–∞–µ–º —Å—Ç—Ä–∞–Ω–∏—Ü—É —Ç–æ–≤–∞—Ä–∞ —Å –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏–µ–º –ø—Ä–æ–∫—Å–∏: {url}")

    response = fetch_with_fallback_proxy(url)
    if not response:
        logger.error(f"–ù–µ —É–¥–∞–ª–æ—Å—å –∑–∞–≥—Ä—É–∑–∏—Ç—å —Å—Ç—Ä–∞–Ω–∏—Ü—É: {url}")
        return {}

    soup = BeautifulSoup(response.content, "html.parser")

    data = {
        "manufacturer": None,
        "article": None,
        "description": None,
        "price": None,
        "analogs": None,
    }

    logger.info("–ò–∑–≤–ª–µ–∫–∞–µ–º –∑–∞–≥–æ–ª–æ–≤–æ–∫ —Ç–æ–≤–∞—Ä–∞ –∫–∞–∫ –æ–ø–∏—Å–∞–Ω–∏–µ")
    title = soup.select_one("h1.product_title.entry-title")
    if title:
        data["description"] = title.text.strip()
        logger.info(f"–û–ø–∏—Å–∞–Ω–∏–µ: {data['description']}")
    else:
        logger.warning("–ó–∞–≥–æ–ª–æ–≤–æ–∫ —Ç–æ–≤–∞—Ä–∞ –Ω–µ –Ω–∞–π–¥–µ–Ω")

    logger.info("–ò–∑–≤–ª–µ–∫–∞–µ–º –∞—Ç—Ä–∏–±—É—Ç—ã (–ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å, –∞—Ä—Ç–∏–∫—É–ª, –∞–Ω–∞–ª–æ–≥–∏)")
    attrs = soup.select("div.wl-attr--list .wl-attr--item")
    for attr in attrs:
        label = attr.get_text(strip=True)
        value = attr.select_one(".pa-right")
        if not value:
            continue
        value = value.text.strip()
        if "–ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å" in label:
            data["manufacturer"] = value
            logger.info(f"–ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å: {value}")
        elif "–ê—Ä—Ç–∏–∫—É–ª" in label:
            data["article"] = value
            logger.info(f"–ê—Ä—Ç–∏–∫—É–ª: {value}")
        elif "–ê–Ω–∞–ª–æ–≥–∏" in label:
            data["analogs"] = value
            logger.info(f"–ê–Ω–∞–ª–æ–≥–∏: {value}")

    logger.info("–ò–∑–≤–ª–µ–∫–∞–µ–º –ø–µ—Ä–≤—É—é –¥–æ—Å—Ç—É–ø–Ω—É—é —Ü–µ–Ω—É")
    price_block = soup.select_one("div.wl-variable--item")
    if price_block:
        price = price_block.select_one(".wl-variable--price")
        if price:
            clean = clean_price(price.text.strip())
            data["price"] = {"price": clean}
            logger.info(f"–¶–µ–Ω–∞: {clean} —Ä—É–±.")
        else:
            logger.warning("–¶–µ–Ω–∞ –Ω–µ –Ω–∞–π–¥–µ–Ω–∞ –≤ –±–ª–æ–∫–µ")
    else:
        logger.warning("–ë–ª–æ–∫ —Å —Ü–µ–Ω–æ–π –Ω–µ –Ω–∞–π–¥–µ–Ω")

    return data

def create_new_excel(filename: str):
    if os.path.exists(filename):
        os.remove(filename)
        print(f"–§–∞–π–ª {filename} —É–¥–∞–ª—ë–Ω")

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append([
        "–ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å",
        "–ê—Ä—Ç–∏–∫—É–ª",
        "–û–ø–∏—Å–∞–Ω–∏–µ",
        "–¶–µ–Ω–∞",
        "–ê–Ω–∞–ª–æ–≥–∏"
    ])
    wb.save(filename)
    print(f"–°–æ–∑–¥–∞–Ω –Ω–æ–≤—ã–π —Ñ–∞–π–ª: {filename}")

def append_products_to_excel(filename: str, products: list[dict]):
    try:
        if not os.path.exists(filename):
            raise FileNotFoundError(f"–§–∞–π–ª {filename} –Ω–µ –Ω–∞–π–¥–µ–Ω. –°–Ω–∞—á–∞–ª–∞ –≤—ã–∑–æ–≤–∏—Ç–µ create_new_excel.")

        wb = load_workbook(filename)
        ws = wb.active

        for product in products:
            try:
                ws.append([
                    product.get("manufacturer", ""),
                    product.get("article", ""),
                    product.get("description", ""),
                    product.get("price", {}).get("price", ""),
                    product.get("analogs", "")
                ])
            except Exception as e:
                logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –¥–æ–±–∞–≤–ª–µ–Ω–∏–∏ —Å—Ç—Ä–æ–∫–∏: {e}")

        wb.save(filename)
        logger.info(f"–î–æ–±–∞–≤–ª–µ–Ω–æ {len(products)} —Å—Ç—Ä–æ–∫ –≤ {filename}")

    except Exception as e:
        logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –∑–∞–ø–∏—Å–∏ –≤ —Ñ–∞–π–ª {filename}: {e}")

def update_config_status(db_connection, name, value):
    """
    –û–±–Ω–æ–≤–ª—è–µ—Ç –∑–Ω–∞—á–µ–Ω–∏–µ –≤ —Ç–∞–±–ª–∏—Ü–µ config.
    –ï—Å–ª–∏ –∑–∞–ø–∏—Å–∏ —Å —É–∫–∞–∑–∞–Ω–Ω—ã–º name –Ω–µ—Ç, –æ–Ω–∞ –±—É–¥–µ—Ç —Å–æ–∑–¥–∞–Ω–∞.
    """
    try:
        if not db_connection.is_connected():
            logging.error("–û—à–∏–±–∫–∞: —Å–æ–µ–¥–∏–Ω–µ–Ω–∏–µ —Å –±–∞–∑–æ–π –¥–∞–Ω–Ω—ã—Ö –æ—Ç—Å—É—Ç—Å—Ç–≤—É–µ—Ç.")
            return
        
        with db_connection.cursor() as cursor:
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —Å—É—â–µ—Å—Ç–≤—É–µ—Ç –ª–∏ –∑–∞–ø–∏—Å—å
            query_check = "SELECT COUNT(*) FROM config WHERE name = %s"
            cursor.execute(query_check, (name,))
            exists = cursor.fetchone()[0]

            if exists:
                # –û–±–Ω–æ–≤–ª—è–µ–º —Å—É—â–µ—Å—Ç–≤—É—é—â—É—é –∑–∞–ø–∏—Å—å
                query_update = "UPDATE config SET value = %s WHERE name = %s"
                cursor.execute(query_update, (value, name))
            else:
                # –°–æ–∑–¥–∞–µ–º –Ω–æ–≤—É—é –∑–∞–ø–∏—Å—å
                query_insert = "INSERT INTO config (name, value) VALUES (%s, %s)"
                cursor.execute(query_insert, (name, value))

            db_connection.commit()
            logging.info(f"–°—Ç–∞—Ç—É—Å '{name}' —É—Å–ø–µ—à–Ω–æ –æ–±–Ω–æ–≤–ª–µ–Ω –¥–æ –∑–Ω–∞—á–µ–Ω–∏—è '{value}'")
    except Exception as e:
        logging.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±–Ω–æ–≤–ª–µ–Ω–∏–∏ —Å—Ç–∞—Ç—É—Å–∞ '{name}' –≤ —Ç–∞–±–ª–∏—Ü–µ config: {e}")
        db_connection.rollback()

def connect_to_db():
    try:
        return mysql.connector.connect(
            host="127.0.0.1",
            user="uploader",
            password="uploader",
            database="avito"
        )
    except mysql.connector.Error as err:
        logging.error(f"–û—à–∏–±–∫–∞ –ø–æ–¥–∫–ª—é—á–µ–Ω–∏—è –∫ –±–∞–∑–µ –¥–∞–Ω–Ω—ã—Ö: {err}")
        raise

if __name__ == "__main__":
    start = time.time()
    try:
        filename = "/home/admin/web/233204.fornex.cloud/public_html/public/products.xlsx"
        create_new_excel(filename)

        links = []
        try:
            db = connect_to_db()
            update_config_status(db, 'parser_status', 'in_progress')
            update_config_status(db, 'parser_update_time', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        except Error:
            logger.error("–û—à–∏–±–∫–∞ –ø—Ä–∏ –ø–æ–¥–∫–ª—é—á–µ–Ω–∏–∏ –∫ –ë–î")

        for page_num in range(1, get_pages_count()):
            try:
                page_url = f"https://trast-zapchast.ru/shop/page/{page_num}/"
                page_links = get_product_links(page_url)
                links.extend(page_links)
                logger.info(f"–°–æ–±—Ä–∞–Ω–æ —Å—Å—ã–ª–æ–∫ —Å {page_url}: {len(page_links)}")
            except Exception as e:
                logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±—Ä–∞–±–æ—Ç–∫–µ —Å—Ç—Ä–∞–Ω–∏—Ü—ã {page_num}: {e}")

        for link in links:
            try:
                product = parse_product_page_single_price(link)
                append_products_to_excel(filename, [product])
            except Exception as e:
                logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±—Ä–∞–±–æ—Ç–∫–µ —Ç–æ–≤–∞—Ä–∞ {link}: {e}")

    except Exception as e:
        logger.critical(f"–§–∞—Ç–∞–ª—å–Ω–∞—è –æ—à–∏–±–∫–∞ –≤—ã–ø–æ–ª–Ω–µ–Ω–∏—è: {e}")
        db = connect_to_db()
        update_config_status(db, 'parser_status', 'failed')
        update_config_status(db, 'parser_update_time', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    try:
        db = connect_to_db()
        update_config_status(db, 'parser_status', 'done')
        update_config_status(db, 'parser_update_time', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    except Error:
        logger.error("–û—à–∏–±–∫–∞ –ø—Ä–∏ –ø–æ–¥–∫–ª—é—á–µ–Ω–∏–∏ –∫ –ë–î")

    end = time.time()
    duration = end - start
    logger.info(f"–ü–∞—Ä—Å–∏–Ω–≥ –∑–∞–≤–µ—Ä—à—ë–Ω –∑–∞ {round(duration, 2)} —Å–µ–∫.")
