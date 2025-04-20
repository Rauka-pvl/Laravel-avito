from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import os
from openpyxl import Workbook, load_workbook
import logging
import requests
import re
import mysql.connector
from mysql.connector import Error
import time
import random
from time import sleep
from concurrent.futures import ThreadPoolExecutor, as_completed

LOGS_PATH = "/home/admin/web/233204.fornex.cloud/public_html/storage/logs/update/"

def setup_logging():
    os.makedirs(LOGS_PATH, exist_ok=True)
    log_filename = os.path.join(LOGS_PATH, f"parser_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    if logger.hasHandlers():
        logger.handlers.clear()
    console_handler = logging.StreamHandler()
    file_handler = logging.FileHandler(log_filename, encoding="utf-8")
    formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    console_handler.setFormatter(formatter)
    file_handler.setFormatter(formatter)
    logger.addHandler(console_handler)
    logger.addHandler(file_handler)
    logging.info(f"Логирование настроено. Файл лога: {log_filename}")
    return log_filename

logger = logging.getLogger(__name__)

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:124.0) Gecko/20100101 Firefox/124.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.6312.107 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_2_1) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.3 Safari/605.1.15",
    "Mozilla/5.0 (Linux; Android 13; Pixel 6 Pro) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Mobile Safari/537.36",
    "Mozilla/5.0 (Linux; Android 12; SAMSUNG SM-G998B) AppleWebKit/537.36 (KHTML, like Gecko) SamsungBrowser/20.0 Chrome/96.0.4664.45 Mobile Safari/537.36",
    "Mozilla/5.0 (Android 13; Mobile; rv:110.0) Gecko/110.0 Firefox/110.0",
    "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Mobile/15E148 Safari/604.1",
    "Mozilla/5.0 (compatible; Googlebot/2.1; +http://www.google.com/bot.html)",
    "Mozilla/5.0 (compatible; bingbot/2.0; +http://www.bing.com/bingbot.htm)"
]

PROXY_LIST = [
    "212.113.123.246:42681",
    "194.87.201.123:24645",
    "92.53.127.107:27807",
    "79.137.133.95:40764",
    "176.124.217.180:12048",
    "92.63.176.36:36162",
    "147.45.254.230:45237",
    "193.164.150.44:57274",
    "89.23.116.215:28305",
    "185.247.185.200:35947",
    "92.255.109.228:25329",
    "176.124.215.172:17789",
    "176.124.216.190:47804",
    "147.45.70.59:47706",
    "proxy.vpnnaruzhu.com:60000:guest2:o7PwR99l",
    "46.232.124.235:50100:bot0TN93:wLiPe9hNN8",
    "149.126.231.118:50100:bot0TN93:wLiPe9hNN8",
    "74.124.221.75:50100:bot0TN93:wLiPe9hNN8",
    "209.200.239.109:50100:bot0TN93:wLiPe9hNN8",
    "45.135.38.65:50100:bot0TN93:wLiPe9hNN8"
]

def get_random_proxy():
    proxy = random.choice(PROXY_LIST)
    if proxy.count(":") == 1:
        return {
            "http": f"http://{proxy}",
            "https": f"http://{proxy}"
        }
    elif proxy.count(":") == 3:
        host, port, user, password = proxy.split(":")
        return {
            "http": f"http://{user}:{password}@{host}:{port}",
            "https": f"http://{user}:{password}@{host}:{port}"
        }
    else:
        raise ValueError(f"Некорректный формат прокси: {proxy}")

def fetch_with_proxy(url: str, max_attempts: int = 10, timeout: int = 10):
    for attempt in range(max_attempts):
        try:
            proxy = get_random_proxy()
            logger.info(f"[{attempt+1}] Пробуем прокси: {proxy['http']}")
            response = requests.get(url, proxies=proxy, timeout=timeout, headers={"User-Agent": random.choice(USER_AGENTS)})
            response.raise_for_status()
            logger.info("✅ Успешное подключение через прокси")
            return response
        except Exception as e:
            logger.warning(f"❌ Ошибка с прокси: {e}")
            sleep(1)
    logger.error("🚫 Не удалось подключиться после всех попыток")
    return None

def fetch_with_fallback_proxy(url: str, timeout: int = 10):
    try:
        logger.info(f"Пробуем загрузить страницу без прокси: {url}")
        response = requests.get(url, timeout=timeout, headers={"User-Agent": random.choice(USER_AGENTS)})
        response.raise_for_status()
        return response
    except Exception as e:
        logger.warning(f"Ошибка без прокси: {e}. Пробуем с прокси...")
        return fetch_with_proxy(url, timeout)

def get_stable_chrome_driver_path():
    os.environ["WDM_LOCAL"] = "1"
    os.environ["WDM_LOG_LEVEL"] = "0"
    os.environ["WDM_CACHE_DIR"] = os.path.join(os.getcwd(), "chrome_driver_cache")
    return ChromeDriverManager().install()

def create_chrome_driver(driver_path):
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--log-level=3")
    options.add_experimental_option("excludeSwitches", ["enable-logging"])
    prefs = {"profile.managed_default_content_settings.images": 2}
    options.add_experimental_option("prefs", prefs)
    return webdriver.Chrome(service=Service(driver_path), options=options)

def create_new_excel(filename):
    if os.path.exists(filename):
        os.remove(filename)
        logger.info(f"Файл {filename} удалён")
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(["Производитель", "Артикул", "Описание", "Цена", "Аналоги"])
    wb.save(filename)
    logger.info(f"Создан новый файл: {filename}")

def append_products_to_excel(filename, products):
    try:
        if not os.path.exists(filename):
            raise FileNotFoundError(f"Файл {filename} не найден.")
        wb = load_workbook(filename)
        ws = wb.active
        for product in products:
            ws.append([
                product.get("manufacturer", ""),
                product.get("article", ""),
                product.get("description", ""),
                product.get("price", {}).get("price", ""),
                product.get("analogs", "")
            ])
            logger.info(f"Добавлен товар: {product.get('description', '')} | Цена: {product.get('price', {}).get('price', '')}")
        wb.save(filename)
        logger.info(f"Добавлено {len(products)} строк в {filename}")
    except Exception as e:
        logger.error(f"Ошибка при записи в файл: {e}")

def get_pages_count(driver, url):
    driver.get(url)
    sleep(2)
    soup = BeautifulSoup(driver.page_source, "html.parser")
    pagination = soup.select_one("div.th-products-view__pagination ul.page-numbers")
    last_page = 1
    if pagination:
        page_links = pagination.find_all("a", class_="page-numbers")
        for link in page_links:
            try:
                page_num = int(link.text.strip())
                last_page = max(last_page, page_num)
            except ValueError:
                continue
    return last_page

def get_product_links(driver, page_url):
    driver.get(page_url)
    sleep(2)
    soup = BeautifulSoup(driver.page_source, "html.parser")
    product_cards = soup.select("li.product")
    product_links = []
    for product in product_cards:
        class_list = product.get("class", [])
        if 'outofstock' in class_list:
            continue
        a_tag = product.find("a", class_="woocommerce-LoopProduct-link")
        href = a_tag.get("href") if a_tag else None
        if href:
            product_links.append(href)
    logger.info(f"Найдено {len(product_links)} ссылок на товары на странице {page_url}")
    return product_links

def clean_price(price_str):
    return re.sub(r"[^\d]", "", price_str)

def parse_product_page_single_price(url: str) -> dict:
    logger.info(f"🔍 Загружаем страницу товара: {url}")

    response = fetch_with_fallback_proxy(url)
    if not response:
        logger.error(f"❌ Не удалось загрузить страницу: {url}")
        return {}

    soup = BeautifulSoup(response.content, "html.parser")

    data = {
        "manufacturer": None,
        "article": None,
        "description": None,
        "price": None,
        "analogs": None,
        "url": url
    }

    # Описание
    title = soup.select_one("h1.product_title.entry-title")
    if title:
        data["description"] = title.text.strip()
    else:
        logger.warning("⚠️ Заголовок товара не найден")

    # Атрибуты
    attrs = soup.select("div.wl-attr--list .wl-attr--item")
    for attr in attrs:
        label = attr.get_text(strip=True)
        value = attr.select_one(".pa-right")
        if not value:
            continue
        value = value.text.strip()
        if "Производитель" in label:
            data["manufacturer"] = value
        elif "Артикул" in label:
            data["article"] = value
        elif "Аналоги" in label:
            data["analogs"] = value

    # Цена
    price_block = soup.select_one("div.wl-variable--item")
    if price_block:
        price = price_block.select_one(".wl-variable--price")
        if price:
            clean = clean_price(price.text.strip())
            data["price"] = {"price": clean}
        else:
            logger.warning("⚠️ Цена не найдена в блоке")
    else:
        logger.warning("⚠️ Блок с ценой не найден")

    logger.info(f"✅ Добавлен товар: {data['description']} | Цена: {data.get('price', {}).get('price', '—')} руб.")
    return data

def connect_to_db(retries=3, delay=3):
    attempt = 0
    while attempt < retries:
        try:
            conn = mysql.connector.connect(
                host="127.0.0.1",
                user="uploader",
                password="uploader",
                database="avito"
            )
            if conn.is_connected():
                logger.info("Успешное подключение к базе данных")
                return conn
        except mysql.connector.Error as err:
            logger.warning(f"Попытка {attempt+1}: ошибка подключения к БД: {err}")
        sleep(delay)
        attempt += 1
    logger.warning("❗ База данных недоступна. Работа продолжается без записи статуса.")
    return None

def update_config_status(db_connection, name, value):
    try:
        with db_connection.cursor() as cursor:
            cursor.execute("SELECT COUNT(*) FROM config WHERE name = %s", (name,))
            exists = cursor.fetchone()[0]
            if exists:
                cursor.execute("UPDATE config SET value = %s WHERE name = %s", (value, name))
            else:
                cursor.execute("INSERT INTO config (name, value) VALUES (%s, %s)", (name, value))
            db_connection.commit()
    except Exception as e:
        db_connection.rollback()
        logger.error(f"Ошибка при обновлении конфигурации: {e}")

if __name__ == "__main__":
    start = time.time()
    setup_logging()
    filename = "/home/admin/web/233204.fornex.cloud/public_html/public/products.xlsx"
    create_new_excel(filename)
    db = connect_to_db()
    if db:
        update_config_status(db, 'parser_status', 'in_progress')
    driver_path = get_stable_chrome_driver_path()
    driver = create_chrome_driver(driver_path)

    try:
        pages = get_pages_count(driver, "https://trast-zapchast.ru/shop/")
        # pages = 5
        all_links = []
        for page_num in range(1, pages + 1):
            page_url = f"https://trast-zapchast.ru/shop/page/{page_num}/"
            all_links.extend(get_product_links(driver, page_url))
    finally:
        driver.quit()

    logger.info(f"Всего получено {len(all_links)} ссылок на товары")
    products = []
    with ThreadPoolExecutor(max_workers=4) as executor:
        future_to_url = {executor.submit(parse_product_page_single_price, url): url for url in all_links}
        for future in as_completed(future_to_url):
            result = future.result()
            if result:
                products.append(result)
                append_products_to_excel(filename, [result])

    if db:
        update_config_status(db, 'parser_status', 'done')
        update_config_status(db, 'parser_update_time', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    total_time = round(time.time() - start, 2)
    logger.info(f"Парсинг завершён за {total_time} секунд. Всего добавлено товаров: {len(products)}")
