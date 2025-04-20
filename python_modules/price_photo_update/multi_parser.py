from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import logging
import requests
import re
import mysql.connector
from mysql.connector import Error
import time
import random
import random
import requests
from time import sleep

LOGS_PATH = "/home/admin/web/233204.fornex.cloud/public_html/storage/logs/update/"
def setup_logging():
    # Уникальное имя файла для каждого запуска
    os.makedirs(LOGS_PATH, exist_ok=True)  # Создаем папку для логов, если ее нет
    log_filename = os.path.join(LOGS_PATH, f"parser_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")

    # Настройка логирования
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)  # Устанавливаем уровень логирования

    # Удаляем старые обработчики (если они есть)
    if logger.hasHandlers():
        logger.handlers.clear()

    # Консольный обработчик
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))

    # Обработчик записи в файл
    file_handler = logging.FileHandler(log_filename, encoding="utf-8")
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))

    # Добавляем обработчики
    logger.addHandler(console_handler)
    logger.addHandler(file_handler)

    # Логируем успешное создание лог-файла
    logging.info(f"Логирование настроено. Файл лога: {log_filename}")

    return log_filename  # Возвращаем имя файла лога


logger = logging.getLogger(__name__)

USER_AGENTS = [
    # Chrome на Windows
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    
    # Firefox на Windows
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:124.0) Gecko/20100101 Firefox/124.0",

    # Chrome на macOS
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.6312.107 Safari/537.36",

    # Safari на macOS
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_2_1) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.3 Safari/605.1.15",

    # Chrome на Android
    "Mozilla/5.0 (Linux; Android 13; Pixel 6 Pro) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Mobile Safari/537.36",

    # Samsung Internet на Android
    "Mozilla/5.0 (Linux; Android 12; SAMSUNG SM-G998B) AppleWebKit/537.36 (KHTML, like Gecko) SamsungBrowser/20.0 Chrome/96.0.4664.45 Mobile Safari/537.36",

    # Firefox на Android
    "Mozilla/5.0 (Android 13; Mobile; rv:110.0) Gecko/110.0 Firefox/110.0",

    # iPhone Safari
    "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Mobile/15E148 Safari/604.1",

    # Googlebot
    "Mozilla/5.0 (compatible; Googlebot/2.1; +http://www.google.com/bot.html)",

    # Bingbot
    "Mozilla/5.0 (compatible; bingbot/2.0; +http://www.bing.com/bingbot.htm)"
]
PROXY_LIST = [
    "212.113.123.246:42681",
    "194.87.201.123:24645",
    "92.53.127.107:27807",
    "79.137.133.95:40764",
    "176.124.217.180:12048",
    "92.63.176.36:36162",
    "147.45.254.230:45237",
    "193.164.150.44:57274",
    "89.23.116.215:28305",
    "185.247.185.200:35947",
    "92.255.109.228:25329",
    "176.124.215.172:17789",
    "176.124.216.190:47804",
    "147.45.70.59:47706",
    "proxy.vpnnaruzhu.com:60000:guest2:o7PwR99l",
    "46.232.124.235:50100:bot0TN93:wLiPe9hNN8",
    "149.126.231.118:50100:bot0TN93:wLiPe9hNN8",
    "74.124.221.75:50100:bot0TN93:wLiPe9hNN8",
    "209.200.239.109:50100:bot0TN93:wLiPe9hNN8",
    "45.135.38.65:50100:bot0TN93:wLiPe9hNN8"
]

def get_random_proxy():
    proxy = random.choice(PROXY_LIST)
    if proxy.count(":") == 1:
        return {
            "http": f"http://{proxy}",
            "https": f"http://{proxy}"
        }
    elif proxy.count(":") == 3:
        host, port, user, password = proxy.split(":")
        return {
            "http": f"http://{user}:{password}@{host}:{port}",
            "https": f"http://{user}:{password}@{host}:{port}"
        }
    else:
        raise ValueError(f"Некорректный формат прокси: {proxy}")

def fetch_with_proxy(url: str, max_attempts: int = 10, timeout: int = 10) -> requests.Response | None:
    for attempt in range(max_attempts):
        try:
            proxy = get_random_proxy()
            print(f"[{attempt+1}] Пробуем прокси: {proxy['http']}")
            response = requests.get(url, proxies=proxy, timeout=timeout)
            response.raise_for_status()
            print("✅ Успешное подключение")
            return response
        except Exception as e:
            print(f"❌ Ошибка с прокси: {e}")
            sleep(1)  # Можно убрать или увеличить паузу
    print("🚫 Не удалось подключиться после всех попыток")
    return None

def fetch_with_fallback_proxy(url: str, timeout: int = 10):
    try:
        logger.info(f"Пробуем загрузить страницу без прокси: {url}")
        response = requests.get(url, timeout=timeout, headers={"User-Agent": random.choice(USER_AGENTS)})
        response.raise_for_status()
        return response
    except Exception as e:
        logger.warning(f"Ошибка без прокси: {e}. Пробуем с прокси...")
        return fetch_with_proxy(url, timeout)

def get_random_headers():
    return {"User-Agent": random.choice(USER_AGENTS)}

def get_pages_count(url: str = "https://trast-zapchast.ru/shop/") -> int:
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    from webdriver_manager.chrome import ChromeDriverManager
    from bs4 import BeautifulSoup
    import time

    logger.info(f"Запуск функции get_pages_count для URL: {url}")

    options = Options()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    try:
        logger.info("Загружаем страницу в браузере")
        driver.get(url)
        time.sleep(3)

        logger.info("Получаем HTML страницы")
        soup = BeautifulSoup(driver.page_source, "html.parser")
        pagination = soup.select_one("div.th-products-view__pagination ul.page-numbers")
        last_page = 1

        if pagination:
            logger.info("Обнаружен блок пагинации, начинаем обработку ссылок")
            page_links = pagination.find_all("a", class_="page-numbers")
            for link in page_links:
                try:
                    page_num = int(link.text.strip())
                    logger.info(f"Обнаружена страница: {page_num}")
                    last_page = max(last_page, page_num)
                except ValueError:
                    logger.info(f"Пропущен элемент пагинации: {link.text.strip()}")
        else:
            logger.warning("Блок пагинации не найден")

        logger.info(f"Общее количество страниц: {last_page}")
        return last_page

    finally:
        driver.quit()
        logger.info("Драйвер закрыт")

def get_product_links(page_url: str) -> list[str]:
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    from webdriver_manager.chrome import ChromeDriverManager
    from bs4 import BeautifulSoup
    import time
    import logging

    logger = logging.getLogger(__name__)
    logger.info(f"Запуск функции get_product_links для страницы: {page_url}")

    options = Options()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    try:
        logger.info("Загружаем страницу в браузере")
        driver.get(page_url)
        time.sleep(3)

        logger.info("Получаем HTML страницы")
        soup = BeautifulSoup(driver.page_source, "html.parser")
        product_links = []

        logger.info("Ищем все карточки товаров")
        product_cards = soup.select("li.product")

        logger.info(f"Обнаружено карточек товаров: {len(product_cards)}")

        for i, product in enumerate(product_cards, 1):
            class_list = product.get("class", [])

            if 'outofstock' in class_list:
                logger.info(f"[{i}] Пропущен — товар не в наличии (class содержит 'outofstock')")
                continue

            a_tag = product.find("a", class_="woocommerce-LoopProduct-link")
            href = a_tag.get("href") if a_tag else None

            if href:
                product_links.append(href)
                logger.info(f"[{i}] Добавлена ссылка: {href}")
            else:
                logger.info(f"[{i}] Ссылка не найдена")

        logger.info(f"Всего добавлено ссылок: {len(product_links)}")
        return product_links

    finally:
        driver.quit()
        logger.info("Драйвер закрыт")

def clean_price(price_str: str) -> str:
    return re.sub(r"[^\d]", "", price_str)

def parse_product_page_single_price(url: str) -> dict:
    logger.info(f"Загружаем страницу товара с использованием прокси: {url}")

    response = fetch_with_fallback_proxy(url)
    if not response:
        logger.error(f"Не удалось загрузить страницу: {url}")
        return {}

    soup = BeautifulSoup(response.content, "html.parser")

    data = {
        "manufacturer": None,
        "article": None,
        "description": None,
        "price": None,
        "analogs": None,
    }

    logger.info("Извлекаем заголовок товара как описание")
    title = soup.select_one("h1.product_title.entry-title")
    if title:
        data["description"] = title.text.strip()
        logger.info(f"Описание: {data['description']}")
    else:
        logger.warning("Заголовок товара не найден")

    logger.info("Извлекаем атрибуты (производитель, артикул, аналоги)")
    attrs = soup.select("div.wl-attr--list .wl-attr--item")
    for attr in attrs:
        label = attr.get_text(strip=True)
        value = attr.select_one(".pa-right")
        if not value:
            continue
        value = value.text.strip()
        if "Производитель" in label:
            data["manufacturer"] = value
            logger.info(f"Производитель: {value}")
        elif "Артикул" in label:
            data["article"] = value
            logger.info(f"Артикул: {value}")
        elif "Аналоги" in label:
            data["analogs"] = value
            logger.info(f"Аналоги: {value}")

    logger.info("Извлекаем первую доступную цену")
    price_block = soup.select_one("div.wl-variable--item")
    if price_block:
        price = price_block.select_one(".wl-variable--price")
        if price:
            clean = clean_price(price.text.strip())
            data["price"] = {"price": clean}
            logger.info(f"Цена: {clean} руб.")
        else:
            logger.warning("Цена не найдена в блоке")
    else:
        logger.warning("Блок с ценой не найден")

    return data

def create_new_excel(filename: str):
    if os.path.exists(filename):
        os.remove(filename)
        print(f"Файл {filename} удалён")

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append([
        "Производитель",
        "Артикул",
        "Описание",
        "Цена",
        "Аналоги"
    ])
    wb.save(filename)
    print(f"Создан новый файл: {filename}")

def append_products_to_excel(filename: str, products: list[dict]):
    try:
        if not os.path.exists(filename):
            raise FileNotFoundError(f"Файл {filename} не найден. Сначала вызовите create_new_excel.")

        wb = load_workbook(filename)
        ws = wb.active

        for product in products:
            try:
                ws.append([
                    product.get("manufacturer", ""),
                    product.get("article", ""),
                    product.get("description", ""),
                    product.get("price", {}).get("price", ""),
                    product.get("analogs", "")
                ])
            except Exception as e:
                logger.error(f"Ошибка при добавлении строки: {e}")

        wb.save(filename)
        logger.info(f"Добавлено {len(products)} строк в {filename}")

    except Exception as e:
        logger.error(f"Ошибка при записи в файл {filename}: {e}")

def connect_to_db(retries=3, delay=3):
    attempt = 0
    while attempt < retries:
        try:
            conn = mysql.connector.connect(
                host="127.0.0.1",
                user="uploader",
                password="uploader",
                database="avito"
            )
            if conn.is_connected():
                logger.info("Успешное подключение к базе данных")
                return conn
        except mysql.connector.Error as err:
            logger.warning(f"Попытка {attempt+1}: ошибка подключения к БД: {err}")
        sleep(delay)
        attempt += 1
    logger.warning("❗ База данных недоступна. Работа продолжается без записи статуса.")
    return None

def update_config_status(db_connection, name, value):
    try:
        with db_connection.cursor() as cursor:
            cursor.execute("SELECT COUNT(*) FROM config WHERE name = %s", (name,))
            exists = cursor.fetchone()[0]
            if exists:
                cursor.execute("UPDATE config SET value = %s WHERE name = %s", (value, name))
            else:
                cursor.execute("INSERT INTO config (name, value) VALUES (%s, %s)", (name, value))
            db_connection.commit()
    except Exception as e:
        db_connection.rollback()
        logger.error(f"Ошибка при обновлении конфигурации: {e}")

if __name__ == "__main__":
    start = time.time()
    setup_logging()
    try:
        filename = "/home/admin/web/233204.fornex.cloud/public_html/public/products.xlsx"
        create_new_excel(filename)

        links = []
        db = connect_to_db()
        if db:
            update_config_status(db, 'parser_status', 'in_progress')
        for page_num in range(1, get_pages_count()):
            try:
                page_url = f"https://trast-zapchast.ru/shop/page/{page_num}/"
                page_links = get_product_links(page_url)
                links.extend(page_links)
                logger.info(f"Собрано ссылок с {page_url}: {len(page_links)}")
            except Exception as e:
                logger.error(f"Ошибка при обработке страницы {page_num}: {e}")

        for link in links:
            try:
                product = parse_product_page_single_price(link)
                append_products_to_excel(filename, [product])
            except Exception as e:
                logger.error(f"Ошибка при обработке товара {link}: {e}")

    except Exception as e:
        logger.critical(f"Фатальная ошибка выполнения: {e}")
        db = connect_to_db()
        update_config_status(db, 'parser_status', 'failed')
        update_config_status(db, 'parser_update_time', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    if db:
        update_config_status(db, 'parser_status', 'done')
        update_config_status(db, 'parser_update_time', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    end = time.time()
    duration = end - start
    logger.info(f"Парсинг завершён за {round(duration, 2)} сек.")
