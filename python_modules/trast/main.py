import os
import re
import time
import random
import logging
import requests
import shutil
import traceback
import queue
import threading
from datetime import datetime
from bs4 import BeautifulSoup
import sys

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
import geckodriver_autoinstaller
import csv
from bz_telebot.database_manager import set_script_start, set_script_end
from proxy_manager import ProxyManager

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
from notification.main import TelegramNotifier

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "avito")))
from config import COMBINED_XML, LOG_DIR, BASE_DIR

LOG_DIR = os.path.join(BASE_DIR, "..", "..", "storage", "app", "public", "output", "logs-trast")
OUTPUT_FILE = os.path.join(LOG_DIR, "..", "trast.xlsx")
TEMP_OUTPUT_FILE = os.path.join(LOG_DIR, "..", "trast_temp.xlsx")
CSV_FILE = os.path.join(LOG_DIR, "..", "trast.csv")
TEMP_CSV_FILE = os.path.join(LOG_DIR, "..", "trast_temp.csv")
BACKUP_FILE = os.path.join(LOG_DIR, "..", "trast_backup.xlsx")
BACKUP_CSV = os.path.join(LOG_DIR, "..", "trast_backup.csv")
os.makedirs(LOG_DIR, exist_ok=True)

logger = logging.getLogger("trast")

# Сохраняем путь к лог-файлу для последующего переименования
LOG_FILE_PATH = os.path.join(LOG_DIR, f"trast_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - [%(threadName)s] - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE_PATH, encoding="utf-8-sig"),
        logging.StreamHandler()
    ]
)

total_products = 0

# Глобальные блокировки для многопоточности
file_lock = threading.Lock()  # Для записи в файлы
stats_lock = threading.Lock()  # Для статистики
webdriver_lock = threading.Lock()  # Для создания WebDriver


class PaginationNotDetectedError(Exception):
    """Поднимается, когда страница каталога выглядит заблокированной и пагинация недоступна."""
    pass


def is_catalog_page_loaded(soup, page_source_lower):
    """
    Эвристика проверки, что страница каталога загрузилась корректно.
    Возвращает True, если найдены явные элементы каталога (пагинация или карточки),
    и отсутствуют признаки заглушек Cloudflare/ошибок.
    """
    has_pagination = bool(soup.select(".facetwp-pager .facetwp-page"))
    has_products = bool(soup.select("div.product.product-plate"))
    
    blocker_keywords = [
        "cloudflare",
        "attention required",
        "checking your browser",
        "just a moment",
        "access denied",
        "forbidden",
        "service temporarily unavailable",
        "temporarily unavailable",
        "maintenance",
        "запрос отклонен",
        "доступ запрещен",
        "ошибка 403",
        "ошибка 503",
        "error 403",
        "error 503",
        "captcha",
        "please enable javascript",
        "varnish cache server",
        "bad gateway",
        "gateway timeout",
    ]
    if any(keyword in page_source_lower for keyword in blocker_keywords):
        return False
    
    # Если на странице нет ни карточек, ни пагинации, считаем что каталог не загрузился
    if not has_products and not has_pagination:
        return False
    
    return True

def has_catalog_structure(soup):
    """
    Проверяет наличие структуры каталога на странице.
    Возвращает True, если найдена хотя бы одна из структурных элементов каталога.
    
    Args:
        soup: BeautifulSoup объект страницы
        
    Returns:
        bool: True если структура каталога присутствует, False иначе
    """
    # Проверяем блок товаров
    has_products_grid = bool(soup.select(".products-grid, .products, .shop-container, .woocommerce-products-header"))
    
    # Проверяем пагинацию
    has_pagination = bool(soup.select(".woocommerce-pagination, .page-numbers, .facetwp-pager, .facetwp-pager .facetwp-page"))
    
    # Проверяем элементы меню/навигации
    has_menu = bool(soup.select("header, .site-header, .main-navigation, nav, .menu, .navigation"))
    
    # Проверяем footer
    has_footer = bool(soup.select("footer, .site-footer, .footer"))
    
    # Проверяем title и meta
    has_title = bool(soup.select("title"))
    has_meta = bool(soup.select("meta"))
    
    # Если есть хотя бы несколько элементов структуры - страница имеет структуру каталога
    structure_count = sum([has_products_grid, has_pagination, has_menu, has_footer, has_title, has_meta])
    
    # Если найдено 3+ элемента структуры - точно есть структура каталога
    # Если найдено 1-2 элемента - возможно частичная загрузка
    return structure_count >= 3

def is_page_blocked(soup, page_source):
    """
    Проверяет, заблокирована ли страница Cloudflare или другими механизмами защиты.
    
    Args:
        soup: BeautifulSoup объект страницы
        page_source: Исходный HTML страницы (строка)
        
    Returns:
        dict: {
            "blocked": bool,
            "reason": str | None
        }
    """
    page_source_lower = page_source.lower() if page_source else ""
    
    # Проверяем ключевые слова блокировки
    blocker_keywords = [
        "cloudflare",
        "attention required",
        "checking your browser",
        "just a moment",
        "access denied",
        "forbidden",
        "service temporarily unavailable",
        "temporarily unavailable",
        "maintenance",
        "запрос отклонен",
        "доступ запрещен",
        "ошибка 403",
        "ошибка 503",
        "error 403",
        "error 503",
        "captcha",
        "please enable javascript",
        "varnish cache server",
        "bad gateway",
        "gateway timeout",
    ]
    
    for keyword in blocker_keywords:
        if keyword in page_source_lower:
            return {
                "blocked": True,
                "reason": keyword
            }
    
    # Проверяем отсутствие структуры каталога
    if not has_catalog_structure(soup):
        return {
            "blocked": True,
            "reason": "no_catalog_structure"
        }
    
    return {
        "blocked": False,
        "reason": None
    }

def is_page_empty(soup, page_source, products_count=0):
    """
    Определяет статус страницы: пустая (конец данных), заблокированная или частично загруженная.
    
    Args:
        soup: BeautifulSoup объект страницы
        page_source: Исходный HTML страницы (строка)
        products_count: Количество найденных товаров на странице
        
    Returns:
        dict: {
            "status": "empty" | "blocked" | "partial",
            "reason": "no_items" | "no_dom" | "cloudflare" | "timeout" | "partial_dom" | "few_items"
        }
    """
    # Сначала проверяем на блокировку
    block_check = is_page_blocked(soup, page_source)
    if block_check["blocked"]:
        return {
            "status": "blocked",
            "reason": block_check["reason"] or "no_dom"
        }
    
    # Проверяем количество товаров
    if products_count == 0:
        # Проверяем наличие структуры каталога
        if has_catalog_structure(soup):
            # Есть структура, но нет товаров - это конец данных
            return {
                "status": "empty",
                "reason": "no_items"
            }
        else:
            # Нет структуры - частичная загрузка или блокировка
            return {
                "status": "partial",
                "reason": "partial_dom"
            }
    elif products_count < 3:
        # Мало товаров (1-2) - подозрение на частичную загрузку или блокировку
        return {
            "status": "partial",
            "reason": "few_items"
        }
    else:
        # Товары есть - страница нормальная
        return {
            "status": "normal",
            "reason": None
        }

def reload_page_if_needed(driver, page_url, max_retries=1):
    """
    Перезагружает страницу если нужно (при частичной загрузке).
    
    Args:
        driver: WebDriver объект
        page_url: URL страницы для загрузки
        max_retries: Максимальное количество попыток перезагрузки
        
    Returns:
        tuple: (soup, products_count) - BeautifulSoup объект и количество товаров
    """
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.by import By
    from selenium.common.exceptions import TimeoutException
    
    for attempt in range(max_retries + 1):
        try:
            if attempt > 0:
                logger.info(f"[RETRY] Перезагружаем страницу {page_url} (попытка {attempt + 1}/{max_retries + 1})...")
                time.sleep(random.uniform(1, 2))
            
            driver.get(page_url)
            time.sleep(random.uniform(3, 6))
            
            # Ждем загрузки страницы
            try:
                wait = WebDriverWait(driver, 15)
                wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
            except TimeoutException:
                logger.warning(f"[WARNING] Таймаут при ожидании загрузки страницы {page_url}")
            
            soup = BeautifulSoup(driver.page_source, "html.parser")
            products = get_products_from_page_soup(soup)
            
            return soup, len(products)
            
        except Exception as e:
            error_msg = str(e).lower()
            # Проверяем, является ли ошибка связанной с прокси
            is_proxy_error = (
                "proxyconnectfailure" in error_msg or
                "proxy" in error_msg and ("refusing" in error_msg or "connection" in error_msg or "failed" in error_msg) or
                "neterror" in error_msg and "proxy" in error_msg
            )
            
            if is_proxy_error:
                # Прокси отказал - не пытаемся дальше перезагружать
                logger.warning(f"[WARNING] Прокси отказал в соединении при перезагрузке страницы (попытка {attempt + 1})")
                # Возвращаем пустой результат, чтобы вызвавший код мог обработать это как ошибку прокси
                return BeautifulSoup("", "html.parser"), 0
            
            logger.warning(f"[WARNING] Ошибка при перезагрузке страницы (попытка {attempt + 1}): {e}")
            if attempt < max_retries:
                continue
            else:
                # Последняя попытка не удалась - возвращаем пустой результат
                return BeautifulSoup("", "html.parser"), 0
    
    return BeautifulSoup("", "html.parser"), 0

def create_new_excel(path):
    if os.path.exists(path):
        os.remove(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(["Manufacturer", "Article", "Description", "Price"])
    wb.save(path)

def create_new_csv(path):
    if os.path.exists(path):
        os.remove(path)
    with open(path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f, delimiter=';')
        writer.writerow(["Manufacturer", "Article", "Description", "Price"])

def append_to_excel(path, product_list):
    """Записывает список товаров в Excel (thread-safe, батч-запись)"""
    global total_products
    if not product_list:
        return
    
    with file_lock:
        if not os.path.exists(path):
            create_new_excel(path)
        try:
            wb = load_workbook(path)
            ws = wb.active
            for p in product_list:
                ws.append([
                    p.get("manufacturer", ""),
                    p.get("article", ""),
                    p.get("description", ""),
                    p.get("price", {}).get("price", "")
                ])
            wb.save(path)
            total_products += len(product_list)
        except Exception as e:
            logger.error(f"Error writing to Excel: {e}")
        try:
            file_size = os.path.getsize(path)
            logger.info(f"Excel updated with {len(product_list)} records, file size: {file_size} bytes")
        except:
            pass

def append_to_csv(path, product_list):
    """Записывает список товаров в CSV (thread-safe, батч-запись)"""
    global total_products
    if not product_list:
        return
    
    with file_lock:
        try:
            with open(path, 'a', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f, delimiter=';')
                for p in product_list:
                    writer.writerow([
                        p.get("manufacturer", ""),
                        p.get("article", ""),
                        p.get("description", ""),
                        p.get("price", {}).get("price", "")
                    ])
            total_products += len(product_list)
        except Exception as e:
            logger.error(f"Error writing to CSV: {e}")

def create_driver(proxy=None, proxy_manager=None, use_chrome=True):
    """Создает Chrome или Firefox драйвер с улучшенным обходом Cloudflare
    
    ВАЖНО: Если прокси SOCKS5/SOCKS4, Chrome автоматически не используется,
    т.к. Chrome не поддерживает SOCKS напрямую.
    """
    # Проверяем тип прокси - если SOCKS, используем Firefox
    if proxy:
        protocol = proxy.get('protocol', 'http').lower()
        if protocol in ['socks4', 'socks5']:
            logger.info(f"Прокси {protocol.upper()} - используем Firefox (Chrome не поддерживает SOCKS)")
            use_chrome = False
    
    # Пробуем сначала Chrome (лучше обходит Cloudflare), потом Firefox
    if use_chrome:
        try:
            return _create_chrome_driver(proxy)
        except (ValueError, Exception) as e:
            # ValueError если SOCKS прокси, другие ошибки - технические проблемы
            if "не поддерживает" in str(e) or "SOCKS" in str(e):
                logger.info(f"Прокси не поддерживается Chrome: {e}, используем Firefox...")
            else:
                logger.warning(f"Chrome не доступен: {e}, пробуем Firefox...")
    
    # Fallback на Firefox
    return _create_firefox_driver(proxy)

def _create_chrome_driver(proxy=None):
    """Создает Chrome драйвер с прокси
    
    Примечание: Chrome НЕ поддерживает SOCKS5 напрямую через --proxy-server.
    Для SOCKS5 прокси используйте Firefox (_create_firefox_driver).
    """
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    from webdriver_manager.chrome import ChromeDriverManager
    from selenium import webdriver
    
    driver_path = ChromeDriverManager().install()
    
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    
    # User-Agent
    user_agents = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    ]
    selected_ua = random.choice(user_agents)
    options.add_argument(f"--user-agent={selected_ua}")
    
    # Настройка прокси для Chrome
    # ВАЖНО: Chrome не поддерживает SOCKS5 напрямую через --proxy-server
    # Используйте только HTTP/HTTPS для Chrome
    if proxy:
        protocol = proxy.get('protocol', 'http').lower()
        ip = proxy['ip']
        port = proxy['port']
        
        if protocol in ['http', 'https']:
            proxy_arg = f"{protocol}://{ip}:{port}"
            options.add_argument(f"--proxy-server={proxy_arg}")
            logger.debug(f"Chrome прокси настроен: {proxy_arg}")
        elif protocol in ['socks4', 'socks5']:
            # SOCKS5 не поддерживается в Chrome напрямую - пропускаем этот прокси
            logger.warning(f"Chrome не поддерживает {protocol.upper()} прокси напрямую. Используйте Firefox для SOCKS прокси.")
            raise ValueError(f"Chrome не поддерживает {protocol.upper()} прокси. Используйте Firefox.")
        else:
            proxy_arg = f"http://{ip}:{port}"
            options.add_argument(f"--proxy-server={proxy_arg}")
            logger.debug(f"Chrome прокси настроен (fallback на HTTP): {proxy_arg}")
    
    service = Service(driver_path)
    driver = webdriver.Chrome(service=service, options=options)
    
    # Stealth скрипты
    stealth_scripts = """
    Object.defineProperty(navigator, 'webdriver', {
        get: () => undefined,
        configurable: true
    });
    window.chrome = { runtime: {} };
    """
    driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {'source': stealth_scripts})
    
    return driver

def _create_firefox_driver(proxy=None):
    """Создает Firefox драйвер с прокси"""
    geckodriver_autoinstaller.install()
    
    options = Options()
    
    # Базовые настройки
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080")
    
    # DNS настройки - НЕ переопределяем при использовании прокси (прокси сам должен делать DNS)
    if not proxy:
        # Только если прокси нет, используем Google DNS
        options.set_preference("network.dns.disablePrefetch", True)
        options.set_preference("network.dns.disablePrefetchFromHTTPS", True)
        options.set_preference("network.dns.defaultIPv4", "8.8.8.8")
        options.set_preference("network.dns.defaultIPv6", "2001:4860:4860::8888")
    else:
        # При прокси - НЕ переопределяем DNS, пусть прокси сам делает DNS резолюцию
        # Для HTTP прокси DNS идет через прокси автоматически
        # Для SOCKS можно включить remote DNS
        if proxy.get('protocol', '').lower() in ['socks4', 'socks5']:
            options.set_preference("network.proxy.socks_remote_dns", True)
    
    # Обход Cloudflare - отключение автоматизации
    options.set_preference("dom.webdriver.enabled", False)
    options.set_preference("useAutomationExtension", False)
    
    # Случайный User-Agent
    user_agents = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/121.0",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:109.0) Gecko/20100101 Firefox/121.0",
        "Mozilla/5.0 (X11; Linux x86_64; rv:109.0) Gecko/20100101 Firefox/121.0",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Edge/120.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.1 Safari/605.1.15"
    ]
    selected_ua = random.choice(user_agents)
    options.set_preference("general.useragent.override", selected_ua)
    
    # Случайные платформы
    platforms = ["Win32", "MacIntel", "Linux x86_64"]
    options.set_preference("general.platform.override", random.choice(platforms))
    
    # Отключение WebRTC для предотвращения утечек IP
    options.set_preference("media.peerconnection.enabled", False)
    options.set_preference("media.navigator.enabled", False)
    
    # Увеличенные таймауты для медленных прокси
    options.set_preference("network.http.connection-timeout", 60)
    options.set_preference("network.http.response.timeout", 60)
    options.set_preference("network.http.keep-alive.timeout", 60)
    options.set_preference("network.http.request.timeout", 60)
    options.set_preference("network.dns.timeout", 30)
    
    # Отключение различных функций и трекинга
    options.set_preference("dom.disable_beforeunload", True)
    options.set_preference("dom.disable_window_open_feature", True)
    options.set_preference("dom.disable_window_move_resize", True)
    options.set_preference("dom.disable_window_flip", True)
    options.set_preference("dom.disable_window_crash_reporter", True)
    
    # Дополнительные настройки обхода детекции
    options.set_preference("privacy.trackingprotection.enabled", True)
    options.set_preference("privacy.trackingprotection.pbmode.enabled", True)
    options.set_preference("browser.safebrowsing.enabled", False)
    options.set_preference("browser.safebrowsing.malware.enabled", False)
    options.set_preference("browser.safebrowsing.phishing.enabled", False)
    options.set_preference("browser.safebrowsing.blockedURIs.enabled", False)
    
    # Отключение автоматических обновлений и телеметрии
    options.set_preference("app.update.enabled", False)
    options.set_preference("app.update.auto", False)
    options.set_preference("toolkit.telemetry.enabled", False)
    options.set_preference("toolkit.telemetry.unified", False)
    options.set_preference("datareporting.healthreport.uploadEnabled", False)
    
    # Настройки SSL/TLS для работы с прокси
    options.set_preference("security.tls.insecure_fallback_hosts", "trast-zapchast.ru")
    options.set_preference("security.tls.unrestricted_rc4_fallback", True)
    options.set_preference("security.tls.version.fallback-limit", 3)
    options.set_preference("security.tls.version.min", 1)
    options.set_preference("security.tls.version.max", 4)
    options.set_preference("security.ssl3.rsa_des_ede3_sha", True)
    options.set_preference("security.ssl3.rsa_rc4_128_sha", True)
    options.set_preference("security.ssl3.rsa_rc4_128_md5", True)
    options.set_preference("security.ssl3.rsa_des_sha", True)
    options.set_preference("security.ssl3.rsa_3des_ede_sha", True)
    options.set_preference("security.ssl3.rsa_aes_128_sha", True)
    options.set_preference("security.ssl3.rsa_aes_256_sha", True)
    options.set_preference("security.ssl3.rsa_aes_128_gcm_sha256", True)
    options.set_preference("security.ssl3.rsa_aes_256_gcm_sha384", True)
    
    # Дополнительные настройки для обхода SSL проблем
    options.set_preference("security.cert_pinning.enforcement_level", 0)
    options.set_preference("security.cert_pinning.process_headers_from_telemetry", False)
    options.set_preference("security.pki.certificate_transparency.mode", 0)
    options.set_preference("security.pki.sha1_enforcement_level", 0)
    
    # Настройка прокси
    if proxy:
        protocol = proxy.get('protocol', 'http').lower()
        ip = proxy['ip']
        port = proxy['port']
        
        if protocol in ['http', 'https']:
            options.set_preference("network.proxy.type", 1)
            options.set_preference("network.proxy.http", ip)
            options.set_preference("network.proxy.http_port", int(port))
            options.set_preference("network.proxy.ssl", ip)
            options.set_preference("network.proxy.ssl_port", int(port))
            options.set_preference("network.proxy.share_proxy_settings", True)
        elif protocol in ['socks4', 'socks5']:
            options.set_preference("network.proxy.type", 1)
            options.set_preference("network.proxy.socks", ip)
            options.set_preference("network.proxy.socks_port", int(port))
            if protocol == 'socks5':
                options.set_preference("network.proxy.socks_version", 5)
            else:
                options.set_preference("network.proxy.socks_version", 4)
            options.set_preference("network.proxy.socks_remote_dns", True)
    
    # Создание драйвера
    service = Service()
    driver = webdriver.Firefox(service=service, options=options)
    
    # Дополнительные скрипты для обхода детекции Cloudflare
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    driver.execute_script("Object.defineProperty(navigator, 'plugins', {get: () => [1, 2, 3, 4, 5]})")
    driver.execute_script("Object.defineProperty(navigator, 'languages', {get: () => ['en-US', 'en']})")
    driver.execute_script("Object.defineProperty(navigator, 'permissions', {get: () => ({query: () => Promise.resolve({state: 'granted'})})})")
    driver.execute_script("Object.defineProperty(navigator, 'hardwareConcurrency', {get: () => 4})")
    driver.execute_script("Object.defineProperty(navigator, 'deviceMemory', {get: () => 8})")
    driver.execute_script("Object.defineProperty(navigator, 'maxTouchPoints', {get: () => 0})")
    
    # Устанавливаем случайные размеры окна
    width = random.randint(1200, 1920)
    height = random.randint(800, 1080)
    driver.set_window_size(width, height)
    
    # Добавляем случайную задержку перед началом работы
    time.sleep(random.uniform(1, 3))
    
    return driver


def get_products_from_page_soup(soup):
    results = []
    cards = soup.select("div.product.product-plate")
    for card in cards:
        stock_badge = card.select_one("div.product-badge.product-stock.instock")
        if not stock_badge or "В наличии" not in stock_badge.text.strip():
            continue

        title_el = card.select_one("a.product-title")
        article_el = card.select_one("div.product-attributes .item:nth-child(1) .value")
        manufacturer_el = card.select_one("div.product-attributes .item:nth-child(2) .value")
        price_el = card.select_one("div.product-price .woocommerce-Price-amount.amount")

        if not (title_el and article_el and manufacturer_el and price_el):
            continue

        title = title_el.text.strip()
        article = article_el.text.strip()
        manufacturer = manufacturer_el.text.strip()
        raw_price = price_el.text.strip().replace("\xa0", " ")
        clean_price = re.sub(r"[^\d\s]", "", raw_price).strip()

        product = {
            "manufacturer": manufacturer,
            "article": article,
            "description": title,
            "price": {"price": clean_price}
        }
        results.append(product)
        logger.info(f"[Product Added] {product}")
    return results

def get_vps_external_ip():
    """
    Получает внешний IP адрес VPS для сравнения.
    Этот IP используется когда прокси НЕ используется.
    
    Returns:
        str: Внешний IP VPS или None если не удалось определить
    """
    # Внешний IP VPS (можно также получить автоматически через requests)
    VPS_EXTERNAL_IP = "31.172.69.102"
    
    # Пробуем несколько сервисов для определения IP VPS
    ip_services = [
        ("https://ifconfig.me/ip", lambda r: r.text.strip()),
    ]
    
    for service_url, extract_func in ip_services:
        try:
            response = requests.get(service_url, timeout=3)
            detected_vps_ip = extract_func(response)
            
            if detected_vps_ip and detected_vps_ip.replace('.', '').isdigit():
                # Проверяем что это похоже на IP адрес
                parts = detected_vps_ip.split('.')
                if len(parts) == 4 and all(0 <= int(p) <= 255 for p in parts if p.isdigit()):
                    # Если автоматически определенный IP совпадает с известным - используем его
                    if detected_vps_ip == VPS_EXTERNAL_IP:
                        logger.debug(f" Внешний IP VPS подтвержден через {service_url}: {VPS_EXTERNAL_IP}")
                        return VPS_EXTERNAL_IP
                    else:
                        # Если IP изменился, логируем предупреждение но используем определенный
                        logger.warning(f"[WARNING]  Внешний IP VPS изменился! Ожидался: {VPS_EXTERNAL_IP}, получен: {detected_vps_ip}")
                        logger.warning(f"   Используем автоматически определенный IP: {detected_vps_ip}")
                        return detected_vps_ip
        except Exception as e:
            logger.debug(f"Не удалось получить IP через {service_url}: {str(e)[:100]}")
            continue
    
    # Если не удалось определить автоматически, используем известный IP
    logger.debug(f"Не удалось автоматически определить IP VPS, используем известный: {VPS_EXTERNAL_IP}")
    return VPS_EXTERNAL_IP

def verify_proxy_usage(driver, proxy):
    """
    Проверяет, что прокси действительно используется через драйвер.
    Пробует несколько сервисов для получения IP и проверяет, что он отличается от локального.
    
    Returns:
        bool: True если прокси используется, False если не удалось подтвердить
    """
    if not proxy:
        return False
    
    proxy_ip = proxy.get('ip', '')
    proxy_country = proxy.get('country', '')
    
    # Получаем внешний IP VPS для сравнения
    vps_external_ip = get_vps_external_ip()
    if vps_external_ip:
        logger.debug(f" Внешний IP VPS (для сравнения): {vps_external_ip}")
    
    # Список сервисов для проверки IP (пробуем несколько для надежности)
    ip_check_services = [
        ("https://ifconfig.me/ip", lambda text: text.strip()),
    ]
    
    external_ips = []
    
    # Сохраняем оригинальный таймаут для восстановления
    original_page_load_timeout = None
    try:
        original_page_load_timeout = driver.timeouts.page_load
    except:
        pass
    
    for service_url, extract_func in ip_check_services:
        try:
            logger.debug(f"Проверка IP через {service_url}...")
            
            # Устанавливаем короткий таймаут для страниц проверки IP
            try:
                driver.set_page_load_timeout(10)  # 10 секунд вместо дефолтных 30+
            except:
                pass
            
            try:
                driver.get(service_url)
            except Exception as timeout_error:
                error_msg = str(timeout_error).lower()
                if "timeout" in error_msg or "timed out" in error_msg:
                    logger.debug(f"  [WARNING]  Таймаут при загрузке {service_url}")
                elif "net::err_" in error_msg:
                    logger.debug(f"  [WARNING]  Сетевая ошибка: {error_msg[:100]}")
                else:
                    logger.debug(f"  [WARNING]  Ошибка загрузки {service_url}: {error_msg[:100]}")
                continue
            
            # Ждем немного для загрузки
            try:
                time.sleep(2)
            except:
                pass
            
            try:
                page_text = driver.page_source.strip()
            except:
                continue
            
            # Проверяем, что это не страница ошибки Chrome
            if "ERR_TIMED_OUT" in page_text or "This site can't be reached" in page_text or "ERR_" in page_text:
                logger.debug(f"  [WARNING]  Страница {service_url} недоступна (страница ошибки Chrome)")
                continue
            
            # Проверяем размер ответа (слишком большие ответы - вероятно HTML ошибки)
            if not page_text or len(page_text) > 200:
                logger.debug(f"  [WARNING]  Неожиданный размер ответа от {service_url}: {len(page_text)} символов")
                continue
            
            try:
                # Всегда ищем IP через regex в page_text (может быть обернут в HTML теги)
                import re
                ip_pattern = r'\b(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})\b'
                
                # Сначала пробуем извлечь через функцию
                raw_ip = extract_func(page_text) if extract_func else page_text
                
                # Ищем IP в raw_ip (может быть HTML)
                ip_matches = re.findall(ip_pattern, str(raw_ip))
                if not ip_matches:
                    # Если не нашли, ищем во всем page_text
                    ip_matches = re.findall(ip_pattern, page_text)
                
                if ip_matches:
                    external_ip = ip_matches[0]  # Берем первый найденный IP
                    logger.debug(f"  Извлечен IP из ответа: {external_ip}")
                else:
                    logger.debug(f"  [WARNING]  Не удалось извлечь IP из ответа")
                    continue
            except Exception as extract_error:
                logger.debug(f"  [WARNING]  Ошибка извлечения IP из ответа {service_url}: {str(extract_error)[:100]}")
                continue
            
            # Проверяем, что это похоже на IP адрес (формат x.x.x.x где x - числа 0-255)
            if external_ip:
                parts = external_ip.split('.')
                if len(parts) == 4:
                    try:
                        # Проверяем что все части - числа в диапазоне 0-255
                        if all(0 <= int(p) <= 255 for p in parts if p.isdigit() and len(p) > 0):
                            external_ips.append(external_ip)
                            logger.info(f"  [OK] IP получен через {service_url}: {external_ip}")
                            break  # Нашли IP, можно не пробовать другие сервисы
                        else:
                            logger.debug(f"  [WARNING]  Неверный формат IP: {external_ip}")
                    except ValueError:
                        logger.debug(f"  [WARNING]  Не удалось распарсить IP: {external_ip}")
                else:
                    logger.debug(f"  [WARNING]  Не похоже на IP (не 4 части): {external_ip[:50]}")
            else:
                logger.debug(f"  [WARNING]  Не удалось извлечь IP из ответа")
        except Exception as e:
            logger.debug(f"  [WARNING]  Неожиданная ошибка при проверке через {service_url}: {str(e)[:100]}")
            continue
    
    # Восстанавливаем оригинальный таймаут
    if original_page_load_timeout is not None:
        try:
            driver.set_page_load_timeout(original_page_load_timeout)
        except:
            try:
                driver.set_page_load_timeout(30)  # Дефолтный таймаут
            except:
                pass
    else:
        try:
            driver.set_page_load_timeout(30)  # Дефолтный таймаут
        except:
            pass
    
    if not external_ips:
        logger.warning("  [WARNING]  Не удалось получить IP ни через один сервис")
        logger.warning("  [WARNING]  Это может быть из-за проблем с сетью, таймаутов или блокировки сервисов проверки IP")
        logger.warning("  [WARNING]  Прокси настроен в драйвере, но не можем подтвердить его использование")
        return False  # Не блокируем работу, но возвращаем False
    
    # Берем первый успешный IP
    detected_ip = external_ips[0]
    
    # Проверяем, что IP действительно через прокси
    # Примечание: IP может не совпадать с IP прокси-сервера, это нормально
    # Главное - проверить, что он не наш локальный IP
    
    logger.info(f" Обнаружен внешний IP через драйвер: {detected_ip}")
    logger.info(f" Прокси: {proxy_ip}:{proxy['port']} ({proxy.get('protocol', 'http').upper()})")
    if proxy_country:
        logger.info(f" Страна прокси: {proxy_country}")
    
    # КРИТИЧЕСКАЯ ПРОВЕРКА: IP должен отличаться от внешнего IP VPS
    if vps_external_ip:
        if detected_ip == vps_external_ip:
            logger.error(f"  [ERROR] ОШИБКА: Обнаружен IP совпадает с внешним IP VPS! Прокси НЕ ИСПОЛЬЗУЕТСЯ!")
            logger.error(f"  Внешний IP VPS: {vps_external_ip}, Обнаружен IP: {detected_ip}")
            logger.error(f"  [WARNING]  Трафик идет напрямую с VPS, без прокси!")
            return False
        else:
            logger.info(f"  [OK] IP отличается от внешнего IP VPS ({vps_external_ip}) - прокси работает!")
            logger.info(f"  [OK] Подтверждено использование прокси: {detected_ip} != {vps_external_ip}")
    
    # Дополнительная проверка: если есть несколько IP от разных сервисов, они должны совпадать
    if len(external_ips) > 1:
        unique_ips = set(external_ips)
        if len(unique_ips) > 1:
            logger.warning(f"  [WARNING]  Разные IP от разных сервисов: {external_ips}")
        else:
            logger.info(f"  [OK] IP подтвержден несколькими сервисами: {detected_ip}")
    
    return True

def get_driver_with_working_proxy_from_list(proxy_manager, proxy_list):
    """Создает драйвер с прокси из списка (пробует Chrome, потом Firefox)"""
    if not proxy_list:
        logger.error("Список прокси пуст")
        return None, 0
    
    for proxy in proxy_list:
        if not proxy:
            continue
            
        protocol = proxy.get('protocol', 'http').lower()
        logger.info(f"Создаем драйвер с прокси {proxy['ip']}:{proxy['port']} ({protocol.upper()})")
        
        driver = None
        can_use_chrome = protocol in ['http', 'https']
        
        if can_use_chrome:
            try:
                logger.info(f"Пробуем этот прокси сначала в Chrome, затем при необходимости в Firefox")
                logger.info(f"  [1/2] Пробуем создать Chrome драйвер с прокси {proxy['ip']}:{proxy['port']}...")
                driver = create_driver(proxy, proxy_manager, use_chrome=True)
                logger.info("[OK] Chrome драйвер создан")
            except Exception as chrome_error:
                logger.warning(f"  [ERROR] Chrome не удалось создать: {str(chrome_error)[:200]}")
                logger.info(f"  [2/2] Пробуем Firefox с тем же прокси {proxy['ip']}:{proxy['port']}...")
                try:
                    driver = create_driver(proxy, proxy_manager, use_chrome=False)
                    logger.info("[OK] Firefox драйвер создан")
                except Exception as firefox_error:
                    logger.error(f"  [ERROR] Firefox тоже не удалось создать: {str(firefox_error)[:200]}")
                    logger.warning(f"[WARNING]  Прокси {proxy['ip']}:{proxy['port']} не работает ни в Chrome, ни в Firefox")
                    continue
        else:
            logger.info(f"Прокси {protocol.upper()} → пропускаем Chrome и сразу используем Firefox")
            try:
                driver = create_driver(proxy, proxy_manager, use_chrome=False)
                logger.info("[OK] Firefox драйвер создан")
            except Exception as firefox_error:
                logger.error(f"  [ERROR] Firefox тоже не удалось создать: {str(firefox_error)[:200]}")
                logger.warning(f"[WARNING]  Прокси {proxy['ip']}:{proxy['port']} не работает ни в Chrome, ни в Firefox")
                continue

        if not driver:
            continue
        
        # ВАЖНО: Проверяем, что прокси действительно используется (неблокирующая проверка)
        try:
            proxy_verified = verify_proxy_usage(driver, proxy)
            if proxy_verified:
                logger.info(f"[OK] ПОДТВЕРЖДЕНО: Прокси {proxy['ip']}:{proxy['port']} используется")
            else:
                logger.warning(f"[WARNING]  Не удалось подтвердить использование прокси через проверку IP")
                logger.warning(f"[WARNING]  Это может быть из-за проблем с сервисами проверки IP или сети")
                logger.warning(f"[WARNING]  Продолжаем работу (прокси настроен в драйвере)")
        except Exception as verify_error:
            logger.warning(f"[WARNING]  Ошибка при проверке прокси: {str(verify_error)[:200]}")
            logger.warning(f"[WARNING]  Продолжаем работу (прокси настроен в драйвере)")
        
        # Сохраняем информацию о прокси в драйвер для последующей проверки
        driver.proxy_info = {
            'ip': proxy['ip'],
            'port': proxy['port'],
            'protocol': proxy.get('protocol', 'http'),
            'country': proxy.get('country', 'Unknown')
        }
        
        # Сохраняем контекст валидации если есть
        validation_context = proxy_manager.validation_cache.get(f"{proxy['ip']}:{proxy['port']}")
        if validation_context:
            driver.trast_validation_context = validation_context
        
        return driver, 0
    
    logger.error("Не удалось создать драйвер ни с одним прокси из списка")
    return None, 0

def get_driver_with_working_proxy(proxy_manager, start_from_index=0):
    """Получает драйвер с рабочим прокси (пробует Chrome, потом Firefox)"""
    max_attempts = 100
    attempt = 0
    
    while attempt < max_attempts:
        try:
            if attempt == 0:
                # Первая попытка - ищем первый рабочий прокси
                proxy = proxy_manager.get_first_working_proxy()
            else:
                # Последующие попытки - ищем следующий рабочий прокси
                proxy, start_from_index = proxy_manager.get_next_working_proxy(start_from_index)
            
            if not proxy:
                logger.error("Не удалось найти рабочий прокси")
                return None, start_from_index
            
            protocol = proxy.get('protocol', 'http').lower()
            logger.info(f"Создаем драйвер с прокси {proxy['ip']}:{proxy['port']} ({protocol.upper()})")
            
            driver = None
            can_use_chrome = protocol in ['http', 'https']
            
            if can_use_chrome:
                try:
                    logger.info(f"Пробуем этот прокси сначала в Chrome, затем при необходимости в Firefox")
                    logger.info(f"  [1/2] Пробуем создать Chrome драйвер с прокси {proxy['ip']}:{proxy['port']}...")
                    with webdriver_lock:
                        driver = create_driver(proxy, proxy_manager, use_chrome=True)
                    logger.info("[OK] Chrome драйвер создан")
                except Exception as chrome_error:
                    logger.warning(f"  [ERROR] Chrome не удалось создать: {str(chrome_error)[:200]}")
                    logger.info(f"  [2/2] Пробуем Firefox с тем же прокси {proxy['ip']}:{proxy['port']}...")
                    try:
                        with webdriver_lock:
                            driver = create_driver(proxy, proxy_manager, use_chrome=False)
                        logger.info("[OK] Firefox драйвер создан")
                    except Exception as firefox_error:
                        logger.error(f"  [ERROR] Firefox тоже не удалось создать: {str(firefox_error)[:200]}")
                        logger.warning(f"[WARNING]  Прокси {proxy['ip']}:{proxy['port']} не работает ни в Chrome, ни в Firefox")
                        logger.info(f"Переходим к следующему прокси...")
                        attempt += 1
                        continue
            else:
                logger.info(f"Прокси {protocol.upper()} → пропускаем Chrome и сразу используем Firefox")
                try:
                    with webdriver_lock:
                        driver = create_driver(proxy, proxy_manager, use_chrome=False)
                    logger.info("[OK] Firefox драйвер создан")
                except Exception as firefox_error:
                    logger.error(f"  [ERROR] Firefox тоже не удалось создать: {str(firefox_error)[:200]}")
                    logger.warning(f"[WARNING]  Прокси {proxy['ip']}:{proxy['port']} не работает ни в Chrome, ни в Firefox")
                    logger.info(f"Переходим к следующему прокси...")
                    attempt += 1
                    continue

            if not driver:
                attempt += 1
                continue
            
            # ВАЖНО: Проверяем, что прокси действительно используется (неблокирующая проверка)
            try:
                proxy_verified = verify_proxy_usage(driver, proxy)
                if proxy_verified:
                    logger.info(f"[OK] ПОДТВЕРЖДЕНО: Прокси {proxy['ip']}:{proxy['port']} используется")
                else:
                    logger.warning(f"[WARNING]  Не удалось подтвердить использование прокси через проверку IP")
                    logger.warning(f"[WARNING]  Это может быть из-за проблем с сервисами проверки IP или сети")
                    logger.warning(f"[WARNING]  Продолжаем работу (прокси настроен в драйвере)")
            except Exception as verify_error:
                logger.warning(f"[WARNING]  Ошибка при проверке прокси: {str(verify_error)[:200]}")
                logger.warning(f"[WARNING]  Продолжаем работу (прокси настроен в драйвере)")
            
            # Сохраняем информацию о прокси в драйвер для последующей проверки
            driver.proxy_info = {
                'ip': proxy['ip'],
                'port': proxy['port'],
                'protocol': proxy.get('protocol', 'http'),
                'country': proxy.get('country', 'Unknown')
            }
            
            # Передаем в драйвер контекст валидации, если он есть
            proxy_key = f"{proxy['ip']}:{proxy['port']}"
            validation_context = proxy_manager.validation_cache.pop(proxy_key, None) if hasattr(proxy_manager, "validation_cache") else None
            if validation_context:
                driver.trast_validation_context = validation_context
                if validation_context.get("total_pages"):
                    logger.info(f"[INFO] Используем кешированное количество страниц: {validation_context['total_pages']}")
            
            return driver, start_from_index
            
        except Exception as e:
            logger.error(f"Ошибка при создании драйвера: {e}")
            attempt += 1
            if attempt < max_attempts:
                logger.info(f"Попытка {attempt + 1}/{max_attempts}")
                time.sleep(2)
    
    # Если все прокси не сработали, ждем 10 минут и обновляем список прокси
    logger.error("Не удалось создать драйвер после всех попыток")
    logger.info("[WAIT] Ожидаем 10 минут перед обновлением списка прокси...")
    logger.info("   Возможно список прокси обновился в репозитории")
    
    wait_minutes = 10
    wait_seconds = wait_minutes * 60
    
    # Показываем прогресс каждую минуту
    for minute in range(wait_minutes):
        remaining = wait_minutes - minute - 1
        logger.info(f"   Осталось ждать: {remaining} минут...")
        time.sleep(60)  # Ждем 1 минуту
    
    logger.info("[OK] Ожидание завершено, обновляем список прокси...")
    
    # Принудительно обновляем список прокси (force_update=True)
    try:
        if proxy_manager.download_proxies(force_update=True):
            logger.info("[OK] Список прокси обновлен, пробуем еще раз...")
            # Сбрасываем индекс и пробуем заново
            return get_driver_with_working_proxy(proxy_manager, start_from_index=0)
        else:
            logger.warning("[WARNING]  Не удалось обновить список прокси, используем кэшированный")
            return None, start_from_index
    except Exception as update_error:
        logger.error(f"[ERROR] Ошибка при обновлении списка прокси: {update_error}")
        logger.error(f"[ERROR] Traceback: {traceback.format_exc()}")
        return None, start_from_index

def get_driver_until_found(proxy_manager, start_from_index=0, max_no_progress_before_reset=20):
    """Получает драйвер с рабочим прокси - ищет до тех пор пока не найдёт
    
    Логика:
    - Ищет рабочий прокси
    - Если не нашли - ждем 1 минуту, обновляем список прокси, продолжаем поиск
    - Повторяем до тех пор пока не найдем
    
    Args:
        proxy_manager: Менеджер прокси
        start_from_index: Индекс для начала поиска следующего прокси
    
    Returns:
        tuple: (driver, start_from_index) - всегда возвращает драйвер (не возвращает None)
    """
    attempt = 0
    last_update_time = time.time()
    update_interval = 60  # Обновляем список прокси каждую минуту
    no_progress_cycles = 0
    
    long_search_thresholds = {50, 100, 200, 300, 400, 500}
    max_attempts_without_success = 500
    
    while True:
        attempt += 1
        logger.info(f"Поиск рабочего прокси (попытка {attempt})...")
        
        if attempt in long_search_thresholds:
            notify_msg = f"[Trast] Уже {attempt} попыток поиска прокси, продолжаем искать..."
            logger.warning(f"[WARNING]  {notify_msg}")
            try:
                TelegramNotifier.notify(notify_msg)
            except Exception as notify_error:
                logger.debug(f"Не удалось отправить уведомление о длительном поиске прокси: {notify_error}")
        
        driver, new_start_from_index = get_driver_with_working_proxy(proxy_manager, start_from_index)
        
        if driver:
            logger.info(f"[OK] Рабочий прокси найден на попытке {attempt}")
            return driver, new_start_from_index
        
        if attempt >= max_attempts_without_success:
            logger.warning(f"[WARNING]  Достигнут лимит в {max_attempts_without_success} попыток поиска прокси без успеха. Принудительно обновляем список и начинаем заново.")
            attempt = 0
            try:
                if proxy_manager.download_proxies(force_update=True):
                    logger.info("[OK] Принудительное обновление списка прокси успешно")
                else:
                    logger.warning("[WARNING]  Принудительное обновление не удалось, используем кэшированный список")
            except Exception as forced_update_error:
                logger.warning(f"[WARNING]  Ошибка при принудительном обновлении прокси: {forced_update_error}")
            proxy_manager.reset_failed_proxies()
            start_from_index = 0
            last_update_time = time.time()
            no_progress_cycles = 0
            continue
        
        # Отслеживаем прогресс по индексу
        if new_start_from_index == start_from_index:
            no_progress_cycles += 1
        else:
            no_progress_cycles = 0
            start_from_index = new_start_from_index
        
        if no_progress_cycles >= max_no_progress_before_reset:
            logger.warning(f"[WARNING]  Не удаётся сделать прогресс по списку прокси ({no_progress_cycles} циклов). Сбрасываем кэш неудачных прокси и начинаем заново.")
            proxy_manager.reset_failed_proxies()
            start_from_index = 0
            no_progress_cycles = 0
        
        # Если не нашли прокси - проверяем, нужно ли обновить список
        current_time = time.time()
        time_since_update = current_time - last_update_time
        
        if time_since_update >= update_interval:
            logger.warning(f"[WAIT] Не найдено рабочего прокси, ждем 1 минуту и обновляем список прокси...")
            logger.info(f"   Ожидание 60 секунд...")
            
            # Показываем прогресс каждые 10 секунд
            for i in range(6):
                remaining = 60 - i * 10
                if remaining > 0:
                    logger.info(f"   Осталось ждать: {remaining} секунд...")
                    time.sleep(10)
            
            logger.info("[OK] Ожидание завершено, обновляем список прокси...")
            
            # Обновляем список прокси
            try:
                if proxy_manager.download_proxies(force_update=True):
                    logger.info("[OK] Список прокси обновлен, продолжаем поиск...")
                    last_update_time = time.time()
                    proxy_manager.reset_failed_proxies()
                    start_from_index = 0  # Сбрасываем индекс после обновления
                else:
                    logger.warning("[WARNING]  Не удалось обновить список прокси, используем кэшированный")
                    last_update_time = time.time()
            except Exception as update_error:
                logger.error(f"[ERROR] Ошибка при обновлении списка прокси: {update_error}")
                last_update_time = time.time()
        else:
            # Если еще не прошла минута - просто ждем немного перед следующей попыткой
            wait_time = 5
            logger.warning(f"[RETRY] Не найдено рабочего прокси, ждем {wait_time} секунд перед следующей попыткой...")
            time.sleep(wait_time)
            start_from_index = new_start_from_index

def get_pages_count_with_driver(driver, url="https://trast-zapchast.ru/shop/"):
    """Получает количество страниц с улучшенной обработкой Cloudflare"""
    try:
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.common.by import By
        
        logger.info("Получаем количество страниц для парсинга...")
        
        # Устанавливаем таймаут для загрузки страницы
        driver.set_page_load_timeout(25)  # 25 секунд - оптимальный баланс
        try:
            driver.get(url)
        except Exception as get_error:
            error_msg = str(get_error).lower()
            if "timeout" in error_msg or "timed out" in error_msg:
                logger.warning(f"[WARNING]  Таймаут при загрузке страницы {url}")
                logger.warning(f"[WARNING]  Прокси может быть слишком медленным или недоступным для целевого сайта")
            raise
        
        # Ждем загрузки страницы и скроллим для активации динамического контента
        wait = WebDriverWait(driver, 30)
        
        # Проверяем на Cloudflare и ждем его прохождения
        page_source_lower = driver.page_source.lower()
        max_wait = 30
        wait_time = 0
        
        while ("cloudflare" in page_source_lower or "checking your browser" in page_source_lower or "just a moment" in page_source_lower) and wait_time < max_wait:
            logger.info(f"[WAIT] Cloudflare проверка... ждем {wait_time}/{max_wait} сек")
            time.sleep(3)
            driver.refresh()
            time.sleep(2)
            page_source_lower = driver.page_source.lower()
            wait_time += 5
        
        # Скроллим для активации динамического контента (как в proxy_manager)
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight/3);")
        time.sleep(random.uniform(1, 2))
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight/2);")
        time.sleep(random.uniform(1, 2))
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(random.uniform(2, 3))
        
        # Дополнительное ожидание для полной загрузки динамического контента
        time.sleep(5)
        
        # Пробуем найти элемент пагинации через Selenium (более надежно для динамического контента)
        total_pages = None
        
        # Метод 1: Ищем через Selenium WebDriverWait (самый надежный для динамического контента)
        try:
            wait = WebDriverWait(driver, 15)
            # Сначала ждем появления пагинации
            pagination_container = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".facetwp-pager")))
            logger.debug("Пагинация найдена через WebDriverWait")
            
            # Пробуем найти элемент .last
            try:
                last_page_element = driver.find_element(By.CSS_SELECTOR, ".facetwp-pager .facetwp-page.last")
                if last_page_element:
                    data_page = last_page_element.get_attribute("data-page")
                    if data_page:
                        total_pages = int(data_page)
                        logger.info(f"[OK] Найдено {total_pages} страниц для парсинга (через Selenium .last)")
                        return total_pages
            except:
                pass
            
            # Если .last не найден, ищем все элементы пагинации и берем максимальный номер
            try:
                page_elements = driver.find_elements(By.CSS_SELECTOR, ".facetwp-pager .facetwp-page")
                logger.debug(f"Найдено элементов пагинации: {len(page_elements)}")
                if page_elements:
                    max_page = 0
                    found_pages = []
                    for page_el in page_elements:
                        data_page = page_el.get_attribute("data-page")
                        if data_page:
                            try:
                                page_num = int(data_page)
                                found_pages.append(page_num)
                                if page_num > max_page:
                                    max_page = page_num
                            except:
                                pass
                    logger.debug(f"Найденные номера страниц: {found_pages}")
                    if max_page > 0:
                        total_pages = max_page
                        logger.info(f"[OK] Найдено {total_pages} страниц для парсинга (через Selenium, максимальный номер из {len(found_pages)} элементов)")
                        return total_pages
                    else:
                        logger.warning(f"[WARNING]  Найдено {len(page_elements)} элементов пагинации, но не удалось извлечь номера страниц")
            except Exception as find_error:
                logger.debug(f"Ошибка при поиске всех элементов пагинации: {find_error}")
        except Exception as wait_error:
            logger.debug(f"WebDriverWait не помог: {wait_error}")
        
        # Метод 2: Пробуем через BeautifulSoup (fallback)
        page_source = driver.page_source
        page_source_lower = page_source.lower()
        soup = BeautifulSoup(page_source, 'html.parser')
        last_page_el = soup.select_one(".facetwp-pager .facetwp-page.last")
        
        if last_page_el and last_page_el.has_attr("data-page"):
            total_pages = int(last_page_el["data-page"])
            logger.info(f"[OK] Найдено {total_pages} страниц для парсинга (через BeautifulSoup .last)")
            return total_pages
        else:
            # Пробуем альтернативные селекторы
            if not last_page_el:
                last_page_el = soup.select_one(".facetwp-page.last")
            if not last_page_el:
                last_page_els = soup.select(".facetwp-pager .facetwp-page")
                logger.debug(f"Найдено элементов пагинации через BeautifulSoup: {len(last_page_els)}")
                if last_page_els:
                    # Берем максимальный номер из всех найденных элементов
                    max_page = 0
                    found_pages = []
                    for page_el in last_page_els:
                        data_page = page_el.get("data-page")
                        if data_page:
                            try:
                                page_num = int(data_page)
                                found_pages.append(page_num)
                                if page_num > max_page:
                                    max_page = page_num
                            except ValueError:
                                continue
                        else:
                            text_value = page_el.get_text(strip=True)
                            if text_value.isdigit():
                                page_num = int(text_value)
                                found_pages.append(page_num)
                                if page_num > max_page:
                                    max_page = page_num
                    logger.debug(f"Найденные номера страниц через BeautifulSoup: {found_pages}")
                    if max_page > 0:
                        total_pages = max_page
                        logger.info(f"[OK] Найдено {total_pages} страниц для парсинга (через BeautifulSoup, максимальный номер из {len(found_pages)} элементов)")
                        return total_pages
                    else:
                        logger.warning(f"[WARNING]  Найдено {len(last_page_els)} элементов пагинации через BeautifulSoup, но не удалось извлечь номера страниц")
                    last_page_el = last_page_els[-1]
            
            if last_page_el and last_page_el.has_attr("data-page"):
                total_pages = int(last_page_el["data-page"])
                logger.info(f"[OK] Найдено {total_pages} страниц для парсинга (альтернативный селектор)")
                return total_pages
            
            has_products = bool(soup.select("div.product.product-plate"))
            has_pagination_any = bool(soup.select(".facetwp-pager .facetwp-page"))
            
            block_indicators = [
                "cloudflare",
                "checking your browser",
                "just a moment",
                "service temporarily unavailable",
                "temporarily unavailable",
                "access denied",
                "ошибка 503",
                "error 503",
                "ошибка 403",
                "error 403",
                "captcha",
                "please enable javascript",
                "attention required",
            ]
            if any(indicator in page_source_lower for indicator in block_indicators):
                logger.warning("[WARNING]  Обнаружены признаки заглушки или блокировки на странице каталога")
                raise PaginationNotDetectedError("Пагинация не найдена из-за блокировки или заглушки")
            
            if not has_products and not has_pagination_any:
                logger.warning("[WARNING]  Каталог не содержит карточек и пагинации — возможно, страница заблокирована")
                raise PaginationNotDetectedError("Пагинация не найдена: отсутствуют карточки и пагинация")
            
            if has_products and not has_pagination_any:
                logger.info("[INFO] Найдены карточки товаров без пагинации — предполагаем одну страницу каталога")
                return 1
            
            # Сохраняем HTML для отладки
            debug_file = os.path.join(LOG_DIR, f"debug_pagination_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html")
            try:
                with open(debug_file, 'w', encoding='utf-8') as f:
                    f.write(page_source)
                logger.warning(f"[WARNING]  Не удалось найти информацию о количестве страниц")
                logger.warning(f"[WARNING]  HTML сохранен в {debug_file} для отладки")
                logger.warning(f"[WARNING]  Размер страницы: {len(page_source)} символов")
                logger.warning(f"[WARNING]  Содержит 'facetwp': {'facetwp' in page_source_lower}")
                logger.warning(f"[WARNING]  Содержит 'shop': {'shop' in page_source_lower}")
            except:
                pass
            
            logger.warning(f"[WARNING]  Используем 1 страницу (не удалось определить количество)")
            return 1
    except Exception as e:
        error_msg = str(e).lower()
        logger.error(f"[ERROR] Ошибка при получении количества страниц: {e}")
        
        # Если это таймаут, не сохраняем HTML (страница не загрузилась)
        if "timeout" in error_msg or "timed out" in error_msg:
            logger.error(f"[ERROR] Таймаут при загрузке страницы - страница не загрузилась")
            logger.error(f"[ERROR] Прокси не может подключиться к целевому сайту или слишком медленный")
        else:
            logger.error(f"[ERROR] Traceback: {traceback.format_exc()}")
            # Сохраняем HTML для отладки только если страница частично загрузилась
            try:
                if driver and hasattr(driver, 'page_source') and driver.page_source:
                    debug_file = os.path.join(LOG_DIR, f"debug_error_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html")
                    with open(debug_file, 'w', encoding='utf-8') as f:
                        f.write(driver.page_source)
                    logger.error(f"[ERROR] HTML сохранен в {debug_file} для отладки")
            except Exception as save_error:
                logger.debug(f"Не удалось сохранить HTML для отладки: {save_error}")
        raise

def worker_thread(thread_id, page_queue, proxy_manager, total_pages, proxy_pool=None):
    """Worker функция для многопоточного парсинга страниц
    
    Args:
        thread_id: ID потока (0, 1, 2)
        page_queue: Очередь страниц для парсинга
        proxy_manager: Менеджер прокси (thread-safe)
        total_pages: Общее количество страниц (для логирования)
        proxy_pool: Пул найденных прокси (thread-safe список)
    """
    # Используем реальное имя потока из threading (будет показано в логах автоматически)
    current_thread = threading.current_thread()
    thread_name = current_thread.name if current_thread.name else f"Thread-{thread_id}"
    logger.info(f"[{thread_name}] === НАЧАЛО РАБОТЫ ПОТОКА {thread_id} ===")
    logger.info(f"[{thread_name}] Поток запущен, получаем прокси...")
    
    # Локальные переменные для потока
    local_buffer = []
    BUFFER_SIZE = 50  # Размер буфера для батч-записи
    empty_pages_count = 0  # Локальный счетчик пустых страниц (для метрики)
    pages_parsed = 0
    products_collected = 0
    
    driver = None
    proxy = None
    
    try:
        # Получаем прокси из пула или через менеджер
        if proxy_pool is not None and len(proxy_pool) > 0:
            # Пытаемся получить прокси из пула (proxy_pool уже является копией, блокировка не нужна)
            if len(proxy_pool) > thread_id:
                proxy = proxy_pool[thread_id]
                logger.info(f"[{thread_name}] Получен прокси из пула: {proxy['ip']}:{proxy['port']} ({proxy.get('protocol', 'http').upper()})")
            else:
                # Если для этого потока нет прокси в пуле, берем первый доступный
                proxy = proxy_pool[0]
                logger.info(f"[{thread_name}] Получен прокси из пула (общий): {proxy['ip']}:{proxy['port']} ({proxy.get('protocol', 'http').upper()})")
        else:
            # Пул пуст или не передан, получаем через менеджер
            if proxy_pool is not None:
                logger.warning(f"[{thread_name}] Пул прокси пуст, получаем через менеджер...")
            proxy = None
        
        # Если не получили из пула, получаем через менеджер
        if not proxy:
            proxy = proxy_manager.get_proxy_for_thread(thread_id)
            if not proxy:
                logger.error(f"[{thread_name}] Не удалось получить прокси для потока")
                return
        
        logger.info(f"[{thread_name}] Получен прокси: {proxy['ip']}:{proxy['port']} ({proxy.get('protocol', 'http').upper()})")
        
        # Создаем драйвер БЕЗ блокировки - каждый поток создает свой драйвер параллельно
        logger.info(f"[{thread_name}] Создаем драйвер (параллельно с другими потоками)...")
        driver = create_driver(proxy, proxy_manager, use_chrome=(proxy.get('protocol', 'http').lower() in ['http', 'https']))
        
        if not driver:
            logger.error(f"[{thread_name}] Не удалось создать драйвер")
            return
        
        logger.info(f"[{thread_name}] Драйвер создан, начинаем парсинг")
        
        # Основной цикл парсинга
        while True:
            try:
                # Получаем страницу из очереди с таймаутом
                try:
                    page_num = page_queue.get(timeout=1)
                except queue.Empty:
                    # Очередь пуста - завершаем поток
                    logger.info(f"[{thread_name}] Очередь пуста, завершаем поток (обработано страниц: {pages_parsed}, товаров: {products_collected})")
                    break
                
                # Парсим страницу
                page_url = f"https://trast-zapchast.ru/shop/?_paged={page_num}"
                logger.info(f"[{thread_name}] 🔄 Начинаем парсинг страницы {page_num}/{total_pages if total_pages else '?'} (очередь: {page_queue.qsize()} страниц)")
                
                try:
                    from selenium.common.exceptions import TimeoutException
                    
                    # Устанавливаем таймаут для загрузки страницы
                    try:
                        driver.set_page_load_timeout(25)
                    except:
                        pass
                    
                    # Загружаем страницу
                    page_load_start = time.time()
                    try:
                        driver.get(page_url)
                        time.sleep(random.uniform(3, 6))
                    except TimeoutException:
                        logger.warning(f"[{thread_name}] Таймаут при загрузке страницы {page_num}")
                        # Пробуем перезагрузить
                        soup, products_count = reload_page_if_needed(driver, page_url, max_retries=1)
                        if products_count == 0:
                            # Проверяем на блокировку
                            page_source = driver.page_source if hasattr(driver, 'page_source') else ""
                            block_check = is_page_blocked(soup, page_source)
                            if block_check["blocked"]:
                                logger.warning(f"[{thread_name}] Page {page_num}: blocked after timeout")
                                # Помечаем страницу как обработанную (не requeue)
                                page_queue.task_done()
                                # Получаем новый прокси
                                proxy = proxy_manager.get_proxy_for_thread(thread_id)
                                if proxy:
                                    try:
                                        driver.quit()
                                    except:
                                        pass
                                    # Создаем драйвер БЕЗ блокировки для параллельной работы
                                    driver = create_driver(proxy, proxy_manager, use_chrome=(proxy.get('protocol', 'http').lower() in ['http', 'https']))
                                    if driver:
                                        logger.info(f"[{thread_name}] Новый прокси получен после таймаута")
                                continue
                    
                    page_load_time = time.time() - page_load_start
                    
                    # Проверяем на блокировку
                    page_source = driver.page_source
                    soup = BeautifulSoup(page_source, "html.parser")
                    block_check = is_page_blocked(soup, page_source)
                    
                    if block_check["blocked"]:
                        logger.warning(f"[{thread_name}] Page {page_num}: blocked → switching proxy")
                        # Получаем новый прокси для потока
                        proxy = proxy_manager.get_proxy_for_thread(thread_id)
                        if proxy:
                            try:
                                driver.quit()
                            except:
                                pass
                            # Создаем драйвер БЕЗ блокировки для параллельной работы
                            driver = create_driver(proxy, proxy_manager, use_chrome=(proxy.get('protocol', 'http').lower() in ['http', 'https']))
                            if driver:
                                logger.info(f"[{thread_name}] Новый прокси получен, продолжаем")
                                page_queue.task_done()
                                continue
                        else:
                            logger.error(f"[{thread_name}] Не удалось получить новый прокси")
                            page_queue.task_done()
                            continue
                    
                    # Получаем товары
                    products = get_products_from_page_soup(soup)
                    products_count = len(products)
                    
                    # Проверяем статус страницы
                    page_status = is_page_empty(soup, page_source, products_count)
                    
                    if page_status["status"] == "normal" and products:
                        # Нормальная страница с товарами
                        local_buffer.extend(products)
                        products_collected += len(products)
                        empty_pages_count = 0
                        
                        # Записываем буфер если он заполнен (только в CSV)
                        if len(local_buffer) >= BUFFER_SIZE:
                            with file_lock:
                                append_to_csv(TEMP_CSV_FILE, local_buffer)
                            logger.info(f"[{thread_name}] Page {page_num}: added {len(products)} products (buffer: {len(local_buffer)}, load time: {page_load_time:.2f}s)")
                            local_buffer.clear()
                    elif page_status["status"] == "empty":
                        # Пустая страница (конец данных) - просто логируем
                        empty_pages_count += 1
                        logger.warning(f"[{thread_name}] Page {page_num}: empty (empty pages: {empty_pages_count})")
                    elif page_status["status"] == "partial":
                        # Частичная загрузка - пробуем перезагрузить
                        logger.warning(f"[{thread_name}] Page {page_num}: partial → retrying")
                        soup, products_count = reload_page_if_needed(driver, page_url, max_retries=1)
                        products = get_products_from_page_soup(soup)
                        if products:
                            local_buffer.extend(products)
                            products_collected += len(products)
                            empty_pages_count = 0
                            
                            if len(local_buffer) >= BUFFER_SIZE:
                                with file_lock:
                                    append_to_csv(TEMP_CSV_FILE, local_buffer)
                                local_buffer.clear()
                    elif page_status["status"] == "blocked":
                        # Блокировка - получаем новый прокси
                        logger.warning(f"[{thread_name}] Page {page_num}: blocked → switching proxy")
                        proxy = proxy_manager.get_proxy_for_thread(thread_id)
                        if proxy:
                            try:
                                driver.quit()
                            except:
                                pass
                            # Создаем драйвер БЕЗ блокировки для параллельной работы
                            driver = create_driver(proxy, proxy_manager, use_chrome=(proxy.get('protocol', 'http').lower() in ['http', 'https']))
                            if driver:
                                logger.info(f"[{thread_name}] Новый прокси получен, продолжаем")
                                page_queue.task_done()
                                continue
                    
                    pages_parsed += 1
                    page_queue.task_done()
                    
                    # Случайная пауза между страницами
                    time.sleep(random.uniform(2, 4))
                    
                except Exception as e:
                    error_msg = str(e).lower()
                    logger.error(f"[{thread_name}] Ошибка при парсинге страницы {page_num}: {e}")
                    
                    # Проверяем на ошибку прокси
                    is_proxy_error = (
                        "proxyconnectfailure" in error_msg or
                        "proxy" in error_msg and ("refusing" in error_msg or "connection" in error_msg or "failed" in error_msg) or
                        "neterror" in error_msg and "proxy" in error_msg
                    )
                    
                    if is_proxy_error:
                        # Пробуем перезагрузить страницу с тем же прокси (один раз)
                        logger.info(f"[{thread_name}] Proxy error on page {page_num}, retrying with same proxy...")
                        try:
                            soup, products_count = reload_page_if_needed(driver, page_url, max_retries=1)
                            products = get_products_from_page_soup(soup)
                            if products:
                                local_buffer.extend(products)
                                products_collected += len(products)
                                if len(local_buffer) >= BUFFER_SIZE:
                                    with file_lock:
                                        append_to_csv(TEMP_CSV_FILE, local_buffer)
                                    local_buffer.clear()
                                page_queue.task_done()
                                pages_parsed += 1
                                continue
                        except Exception as retry_error:
                            logger.warning(f"[{thread_name}] Retry failed: {retry_error}")
                        
                        # Если retry не помог - получаем новый прокси
                        logger.info(f"[{thread_name}] Getting new proxy after proxy error...")
                        proxy = proxy_manager.get_proxy_for_thread(thread_id)
                        if proxy:
                            try:
                                driver.quit()
                            except:
                                pass
                            # Создаем драйвер БЕЗ блокировки для параллельной работы
                            driver = create_driver(proxy, proxy_manager, use_chrome=(proxy.get('protocol', 'http').lower() in ['http', 'https']))
                            if driver:
                                logger.info(f"[{thread_name}] Новый прокси получен после ошибки")
                    
                    page_queue.task_done()
                    
            except Exception as e:
                logger.error(f"[{thread_name}] Критическая ошибка в цикле парсинга: {e}")
                page_queue.task_done()
                continue
        
        # Записываем оставшиеся товары из буфера (только в CSV)
        if local_buffer:
            with file_lock:
                append_to_csv(TEMP_CSV_FILE, local_buffer)
            logger.info(f"[{thread_name}] Записаны оставшиеся товары из буфера: {len(local_buffer)}")
        
        logger.info(f"[{thread_name}] === ЗАВЕРШЕНИЕ ПОТОКА {thread_id} ===")
        logger.info(f"[{thread_name}] Статистика: страниц обработано: {pages_parsed}, товаров собрано: {products_collected}")
        
    except Exception as e:
        logger.error(f"[{thread_name}] Критическая ошибка в worker: {e}")
        logger.error(f"[{thread_name}] Traceback: {traceback.format_exc()}")
    finally:
        # Закрываем драйвер
        if driver:
            try:
                driver.quit()
            except:
                pass

def producer(proxy_manager):
    """Основная функция парсинга ТОЛЬКО через прокси
    
    Логика:
    - Запускаем многопоточный поиск 3 прокси параллельно
    - Как только найден первый прокси - сразу запускаем парсинг
    - Остальные потоки продолжают искать прокси в фоне
    - Каждому воркеру парсинга выдаем готовый прокси из пула
    - При блокировке/ошибке: запоминаем страницу, ищем новый прокси до тех пор пока не найдем
    - Продолжаем парсинг с запомненной страницы
    - Останавливаемся при 2 пустых страницах подряд (конец данных)
    - Различаем пустую страницу (конец данных) от блокировки (нет структуры DOM)
    """
    thread_name = "MainThread"
    logger.info(f"[{thread_name}] Starting producer with PROXY-ONLY strategy (multithreaded proxy search)")
    
    # Запускаем многопоточный поиск 3 прокси параллельно
    logger.info(f"[{thread_name}] === ЗАПУСК МНОГОПОТОЧНОГО ПОИСКА ПРОКСИ ===")
    logger.info(f"[{thread_name}] Ищем 3 рабочих прокси в 3 потоках...")
    
    # Запускаем поиск прокси в отдельном потоке, чтобы не блокировать
    # Используем список внутри proxy_manager для thread-safe доступа
    found_proxies_list = []
    proxy_search_thread = None
    first_proxy_ready = threading.Event()
    
    def search_proxies_background():
        """Фоновая функция поиска прокси"""
        nonlocal found_proxies_list
        try:
            proxies = proxy_manager.get_working_proxies_parallel(count=3, max_attempts_per_thread=50)
            # Thread-safe добавление в список
            with proxy_manager.lock:
                found_proxies_list.extend(proxies)
            if proxies:
                logger.info(f"[{thread_name}] Многопоточный поиск завершен: найдено {len(proxies)} прокси")
                first_proxy_ready.set()
            else:
                logger.warning(f"[{thread_name}] Многопоточный поиск не нашел рабочих прокси")
        except Exception as e:
            logger.error(f"[{thread_name}] Ошибка в фоновом поиске прокси: {e}")
    
    # Запускаем поиск прокси в фоне
    proxy_search_thread = threading.Thread(target=search_proxies_background, daemon=False, name="ProxySearch-Background")
    proxy_search_thread.start()
    logger.info(f"[{thread_name}] Фоновый поиск прокси запущен")
    
    # Ждем первого прокси (с таймаутом)
    logger.info(f"[{thread_name}] Ожидаем первый рабочий прокси...")
    wait_timeout = 300  # 5 минут максимум
    start_wait = time.time()
    
    # Периодически проверяем, найден ли первый прокси
    while not first_proxy_ready.is_set() and (time.time() - start_wait) < wait_timeout:
        time.sleep(1)
        # Проверяем, есть ли уже найденные прокси (thread-safe)
        with proxy_manager.lock:
            if found_proxies_list:
                first_proxy_ready.set()
                break
    
    # Получаем найденные прокси (thread-safe)
    with proxy_manager.lock:
        found_proxies = found_proxies_list.copy()
    
    if not found_proxies:
        logger.error(f"[{thread_name}] Не удалось найти рабочий прокси за {wait_timeout} секунд")
        # Ждем завершения фонового потока
        if proxy_search_thread:
            proxy_search_thread.join(timeout=10)
        return 0, {"pages_checked": 0, "proxy_switches": 0, "cloudflare_blocks": 0, "max_empty_streak": 0}
    
    logger.info(f"[{thread_name}] [OK] Найден первый рабочий прокси: {found_proxies[0]['ip']}:{found_proxies[0]['port']}")
    logger.info(f"[{thread_name}] Всего найдено прокси: {len(found_proxies)} (поиск продолжается в фоне)")
    
    # Получаем драйвер с первым найденным прокси для получения total_pages
    proxy = found_proxies[0]
    logger.info(f"[{thread_name}] Создаем драйвер с первым прокси для получения количества страниц...")
    driver, start_from_index = get_driver_with_working_proxy_from_list(proxy_manager, [proxy])
    if not driver:
        logger.error(f"[{thread_name}] Не удалось создать драйвер с первым прокси")
        return 0, {"pages_checked": 0, "proxy_switches": 0, "cloudflare_blocks": 0, "max_empty_streak": 0}
    
    logger.info(f"[{thread_name}] [OK] Драйвер создан, получаем количество страниц")
    
    total_collected = 0
    empty_pages_count = 0
    max_empty_pages = 2  # Изменено с 3 на 2 - оптимальный баланс
    pages_checked = 0  # Счетчик проверенных страниц
    max_empty_streak = 0
    proxy_switch_count = 0
    cloudflare_block_count = 0
    forced_proxy_updates = 0
    
    # Защита от бесконечной смены прокси
    proxy_switch_times = []  # Список времен смены прокси
    max_proxy_switches_per_period = 20  # Максимум смен прокси за период
    proxy_switch_period_seconds = 600  # 10 минут
    
    try:
        logger.info(f"Начинаем парсинг ТОЛЬКО через прокси")
        
        # Получаем количество страниц (безлимитные повторные попытки)
        total_pages = None
        last_page_count_error = None
        cached_context = getattr(driver, "trast_validation_context", None)
        cached_total_pages = None
        if cached_context:
            cached_total_pages = cached_context.get("total_pages")
        
        fallback_mode = False
        fallback_threshold = 10
        proxy_refresh_every_failures = 5
        page_count_attempts = 0
        page_count_failures = 0
        
        while True:
            use_cached_value = cached_total_pages and cached_total_pages > 0 and page_count_attempts == 0
            page_count_attempts += 1
            
            try:
                if use_cached_value:
                    logger.info(f"Попытка #{page_count_attempts} получить количество страниц... [используем кеш]")
                    total_pages = cached_total_pages
                    cached_total_pages = None  # Кеш используем только один раз
                else:
                    logger.info(f"Попытка #{page_count_attempts} получить количество страниц...")
                    total_pages = get_pages_count_with_driver(driver)
                
                if total_pages and total_pages > 0:
                    logger.info(f"[OK] Успешно получено количество страниц: {total_pages}")
                    break
                
                logger.warning(f"[WARNING]  Получено некорректное количество страниц: {total_pages}")
                total_pages = None
                page_count_failures += 1
            except Exception as e:
                last_page_count_error = e
                page_count_failures += 1
                if isinstance(e, PaginationNotDetectedError):
                    logger.warning(f"[WARNING]  Пагинация не определена (попытка #{page_count_attempts}): {e}")
                else:
                    logger.error(f"[ERROR] Ошибка при получении количества страниц (попытка #{page_count_attempts}): {e}")
                total_pages = None
            
            if total_pages and total_pages > 0:
                break
            
            stats = proxy_manager.get_proxy_queue_stats()
            logger.info(
                "Статус очереди прокси: всего=%s, доступно=%s, успешных=%s, исключено=%s",
                stats.get("total"),
                stats.get("available"),
                stats.get("successful"),
                stats.get("failed"),
            )
            
            if page_count_failures > 0 and page_count_failures % proxy_refresh_every_failures == 0:
                logger.warning(f"[WARNING]  {page_count_failures} неудачных попыток получить количество страниц — принудительно обновляем список прокси")
                try:
                    if proxy_manager.download_proxies(force_update=True):
                        logger.info("[OK] Список прокси обновлен (page_count)")
                        proxy_manager.reset_failed_proxies()
                        start_from_index = 0
                    else:
                        logger.warning("[WARNING]  Не удалось обновить список прокси принудительно (page_count)")
                except Exception as refresh_error:
                    logger.warning(f"[WARNING]  Ошибка при принудительном обновлении списка прокси: {refresh_error}")
            
            if page_count_attempts >= fallback_threshold:
                logger.warning("[WARNING]  Не удалось определить количество страниц после множества попыток — переходим в fallback режим без total_pages")
                fallback_mode = True
                break
            
            logger.info("Ищем новый прокси для повторной попытки получить количество страниц...")
            try:
                driver.quit()
            except Exception:
                pass
            
            driver, start_from_index = get_driver_until_found(proxy_manager, start_from_index)
            proxy_switch_count += 1
            forced_proxy_updates += 1
            
            wait_before_retry = random.uniform(5, 10)
            logger.debug(f"Ждем {wait_before_retry:.1f} секунд перед следующей попыткой")
            time.sleep(wait_before_retry)
        
        # Закрываем драйвер, который использовался для получения total_pages
        try:
            driver.quit()
        except:
            pass
        
        # Многопоточный режим: создаем очередь страниц и запускаем 3 потока
        if total_pages and total_pages > 0:
            logger.info(f"[{thread_name}] Запускаем многопоточный парсинг: {total_pages} страниц в 3 потоках")
            
            # Создаем очередь страниц
            page_queue = queue.Queue()
            for page in range(1, total_pages + 1):
                page_queue.put(page)
            
            logger.info(f"[{thread_name}] Очередь создана: {total_pages} страниц")
            
            # Ждем завершения поиска прокси или используем уже найденные
            if proxy_search_thread and proxy_search_thread.is_alive():
                # Получаем текущий список (thread-safe)
                with proxy_manager.lock:
                    current_count = len(found_proxies_list)
                logger.info(f"[{thread_name}] Ожидаем завершения поиска прокси (найдено: {current_count})...")
                # Ждем максимум 60 секунд для завершения поиска
                proxy_search_thread.join(timeout=60)
            
            # Обновляем список найденных прокси (thread-safe)
            with proxy_manager.lock:
                current_found_proxies = found_proxies_list.copy()
            
            logger.info(f"[{thread_name}] Пул прокси для парсинга: {len(current_found_proxies)} прокси")
            for i, p in enumerate(current_found_proxies):
                logger.info(f"[{thread_name}]   Прокси {i}: {p['ip']}:{p['port']} ({p.get('protocol', 'http').upper()})")
            
            # Запускаем 3 потока ПАРАЛЛЕЛЬНО с пулом прокси
            threads = []
            logger.info(f"[{thread_name}] === ЗАПУСК МНОГОПОТОЧНОГО ПАРСИНГА ===")
            for thread_id in range(3):
                thread = threading.Thread(
                    target=worker_thread,
                    args=(thread_id, page_queue, proxy_manager, total_pages, current_found_proxies),
                    daemon=False,
                    name=f"Worker-{thread_id}"  # Явно задаем имя потока для логирования
                )
                thread.start()
                threads.append(thread)
                logger.info(f"[{thread_name}] ✓ Поток {thread_id} запущен (имя: Worker-{thread_id})")
                # Небольшая задержка между запусками, чтобы видеть параллельность
                time.sleep(0.1)
            
            logger.info(f"[{thread_name}] Все 3 потока запущены и работают ПАРАЛЛЕЛЬНО")
            logger.info(f"[{thread_name}] Ожидаем завершения всех потоков...")
            
            # Ждем завершения всех потоков
            for i, thread in enumerate(threads):
                thread.join()
                logger.info(f"[{thread_name}] Поток {i} завершен")
            
            logger.info(f"[{thread_name}] === ВСЕ ПОТОКИ ЗАВЕРШЕНЫ ===")
            
            # Ждем завершения фонового поиска прокси (если еще не завершился)
            if proxy_search_thread and proxy_search_thread.is_alive():
                logger.info(f"[{thread_name}] Ожидаем завершения фонового поиска прокси...")
                proxy_search_thread.join(timeout=30)
            
            # Возвращаем метрики (упрощенные, т.к. детальная статистика собирается в потоках)
            return total_products, {
                "pages_checked": total_pages,
                "proxy_switches": 0,  # Собирается в потоках
                "cloudflare_blocks": 0,  # Собирается в потоках
                "max_empty_streak": 0,  # Собирается в потоках
            }
        
        # Fallback режим (без total_pages) - используем старую логику
        logger.warning(f"[{thread_name}] Fallback режим: парсинг без total_pages (однопоточный)")
        current_page = 1
        
        while True:
            if not fallback_mode and total_pages and current_page > total_pages:
                break
            
            try:
                from selenium.common.exceptions import TimeoutException
                
                page_url = f"https://trast-zapchast.ru/shop/?_paged={current_page}"
                if fallback_mode or not total_pages:
                    logger.info(f"[{thread_name}] Parsing page {current_page} (проверено: {pages_checked}) [fallback без total_pages]")
                else:
                    logger.info(f"[{thread_name}] Parsing page {current_page}/{total_pages} (проверено: {pages_checked})")
                
                # Устанавливаем таймаут для загрузки страницы
                try:
                    driver.set_page_load_timeout(25)
                except:
                    pass
                
                page_load_start = time.time()
                try:
                    driver.get(page_url)
                    time.sleep(random.uniform(3, 6))
                except TimeoutException:
                    logger.warning(f"[WARNING] Таймаут при загрузке страницы {current_page} (25 сек)")
                    # Пробуем перезагрузить страницу
                    soup, products_count = reload_page_if_needed(driver, page_url, max_retries=1)
                    if products_count == 0:
                        # Страница не загрузилась - проверяем на блокировку
                        page_source = driver.page_source if hasattr(driver, 'page_source') else ""
                        block_check = is_page_blocked(soup, page_source)
                        if block_check["blocked"]:
                            logger.warning(f"[BLOCKED] Page {current_page}: DOM missing → proxy blocked → switching proxy")
                            blocked_page = current_page
                            cloudflare_block_count += 1
                            # Проверяем защиту от бесконечной смены прокси
                            current_time = time.time()
                            proxy_switch_times[:] = [t for t in proxy_switch_times if current_time - t < proxy_switch_period_seconds]
                            if len(proxy_switch_times) >= max_proxy_switches_per_period:
                                logger.error(f"[ERROR] Превышен лимит смен прокси ({max_proxy_switches_per_period} за {proxy_switch_period_seconds} сек). Аварийное завершение.")
                                break
                            
                            try:
                                driver.quit()
                            except:
                                pass
                            driver, start_from_index = get_driver_until_found(proxy_manager, start_from_index)
                            proxy_switch_count += 1
                            proxy_switch_times.append(time.time())
                            logger.info(f"[OK] Новый прокси найден, продолжаем парсинг со страницы {blocked_page}")
                            continue
                        else:
                            # Частичная загрузка - продолжаем
                            logger.warning(f"[PARTIAL] Page {current_page}: DOM incomplete → retrying load")
                            continue
                
                page_load_time = time.time() - page_load_start
                
                # Проверяем на блокировку используя новые функции
                page_source = driver.page_source
                soup = BeautifulSoup(page_source, "html.parser")
                block_check = is_page_blocked(soup, page_source)
                
                if block_check["blocked"]:
                    logger.warning(f"[BLOCKED] Page {current_page}: DOM missing → proxy blocked → switching proxy (reason: {block_check['reason']})")
                    # Запоминаем текущую страницу
                    blocked_page = current_page
                    cloudflare_block_count += 1
                    
                    # Проверяем защиту от бесконечной смены прокси
                    current_time = time.time()
                    proxy_switch_times[:] = [t for t in proxy_switch_times if current_time - t < proxy_switch_period_seconds]
                    if len(proxy_switch_times) >= max_proxy_switches_per_period:
                        logger.error(f"[ERROR] Превышен лимит смен прокси ({max_proxy_switches_per_period} за {proxy_switch_period_seconds} сек). Аварийное завершение.")
                        break
                    
                    try:
                        driver.quit()
                    except:
                        pass
                    # Ищем новый прокси до тех пор пока не найдем
                    logger.info(f"Ищем новый рабочий прокси для продолжения парсинга со страницы {blocked_page}...")
                    driver, start_from_index = get_driver_until_found(proxy_manager, start_from_index)
                    proxy_switch_count += 1
                    proxy_switch_times.append(time.time())
                    logger.info(f"[OK] Новый прокси найден, продолжаем парсинг со страницы {blocked_page}")
                    continue
                
                # Получаем товары
                products = get_products_from_page_soup(soup)
                products_count = len(products)
                
                # Проверяем статус страницы
                page_status = is_page_empty(soup, page_source, products_count)
                
                if page_status["status"] == "normal" and products:
                    # Пишем во временные файлы (старый файл не трогаем)
                    append_to_csv(TEMP_CSV_FILE, products)
                    append_to_csv(TEMP_CSV_FILE, products)
                    logger.info(f"[{thread_name}] Page {current_page}: added {len(products)} products (load time: {page_load_time:.2f}s)")
                    total_collected += len(products)
                    empty_pages_count = 0  # Сбрасываем счетчик пустых страниц только при успешном получении товаров
                elif page_status["status"] == "partial":
                    # Частичная загрузка - повторяем попытку
                    logger.warning(f"[PARTIAL] Page {current_page}: DOM incomplete → retrying load (reason: {page_status['reason']})")
                    soup, products_count = reload_page_if_needed(driver, page_url, max_retries=1)
                    products = get_products_from_page_soup(soup)
                    if products:
                        append_to_csv(TEMP_CSV_FILE, products)
                        append_to_csv(TEMP_CSV_FILE, products)
                        logger.info(f"[{thread_name}] Page {current_page} (retry): added {len(products)} products")
                        total_collected += len(products)
                        empty_pages_count = 0
                    else:
                        # После повторной попытки все еще нет товаров - проверяем статус
                        page_status = is_page_empty(soup, page_source, 0)
                        if page_status["status"] == "empty":
                            empty_pages_count += 1
                            if empty_pages_count > max_empty_streak:
                                max_empty_streak = empty_pages_count
                            logger.warning(f"[EMPTY PAGE] Page {current_page}: no items but catalog structure exists → end of data logic triggered (empty pages: {empty_pages_count})")
                        elif page_status["status"] == "blocked":
                            # Блокировка после повторной попытки
                            logger.warning(f"[BLOCKED] Page {current_page}: DOM missing after retry → proxy blocked → switching proxy")
                            blocked_page = current_page
                            cloudflare_block_count += 1
                            
                            # Проверяем защиту от бесконечной смены прокси
                            current_time = time.time()
                            proxy_switch_times[:] = [t for t in proxy_switch_times if current_time - t < proxy_switch_period_seconds]
                            if len(proxy_switch_times) >= max_proxy_switches_per_period:
                                logger.error(f"[ERROR] Превышен лимит смен прокси ({max_proxy_switches_per_period} за {proxy_switch_period_seconds} сек). Аварийное завершение.")
                                break
                            
                            try:
                                driver.quit()
                            except:
                                pass
                            driver, start_from_index = get_driver_until_found(proxy_manager, start_from_index)
                            proxy_switch_count += 1
                            proxy_switch_times.append(time.time())
                            logger.info(f"[OK] Новый прокси найден, продолжаем парсинг со страницы {blocked_page}")
                            continue
                elif page_status["status"] == "empty":
                    # Пустая страница (конец данных) - НЕ меняем прокси
                    empty_pages_count += 1
                    if empty_pages_count > max_empty_streak:
                        max_empty_streak = empty_pages_count
                    logger.warning(f"[EMPTY PAGE] Page {current_page}: no items but catalog structure exists → end of data logic triggered (empty pages: {empty_pages_count})")
                    
                    # Условие остановки: если 2 страницы подряд пустые (конец данных)
                    if empty_pages_count >= max_empty_pages:
                        logger.info(f"[EMPTY PAGE] Найдено {max_empty_pages} пустых страниц подряд. Останавливаем парсинг.")
                        pages_checked += 1
                        break
                elif page_status["status"] == "blocked":
                    # Блокировка - меняем прокси
                    logger.warning(f"[BLOCKED] Page {current_page}: DOM missing → proxy blocked → switching proxy (reason: {page_status['reason']})")
                    blocked_page = current_page
                    cloudflare_block_count += 1
                    
                    # Проверяем защиту от бесконечной смены прокси
                    current_time = time.time()
                    proxy_switch_times[:] = [t for t in proxy_switch_times if current_time - t < proxy_switch_period_seconds]
                    if len(proxy_switch_times) >= max_proxy_switches_per_period:
                        logger.error(f"[ERROR] Превышен лимит смен прокси ({max_proxy_switches_per_period} за {proxy_switch_period_seconds} сек). Аварийное завершение.")
                        break
                    
                    try:
                        driver.quit()
                    except:
                        pass
                    driver, start_from_index = get_driver_until_found(proxy_manager, start_from_index)
                    proxy_switch_count += 1
                    proxy_switch_times.append(time.time())
                    logger.info(f"[OK] Новый прокси найден, продолжаем парсинг со страницы {blocked_page}")
                    continue
                
                pages_checked += 1
                current_page += 1
                
                # Случайная пауза между страницами
                time.sleep(random.uniform(2, 4))
                
            except TimeoutException as e:
                logger.error(f"[ERROR] TimeoutException при парсинге страницы {current_page}: {e}")
                # Таймаут - пробуем перезагрузить страницу
                logger.info(f"[RETRY] Повторная попытка загрузки страницы {current_page} после таймаута...")
                try:
                    soup, products_count = reload_page_if_needed(driver, page_url, max_retries=1)
                    products_retry = get_products_from_page_soup(soup)
                    if products_retry:
                        append_to_csv(TEMP_CSV_FILE, products_retry)
                        append_to_csv(TEMP_CSV_FILE, products_retry)
                        logger.info(f"[{thread_name}] Page {current_page} (retry after timeout): added {len(products_retry)} products")
                        total_collected += len(products_retry)
                        empty_pages_count = 0
                        pages_checked += 1
                        current_page += 1
                        time.sleep(random.uniform(2, 4))
                        continue
                except Exception as retry_error:
                    logger.warning(f"[WARNING] Повторная попытка страницы {current_page} после таймаута не удалась: {retry_error}")
                
                # Если повторная попытка не удалась - проверяем на блокировку и меняем прокси
                errored_page = current_page
                try:
                    page_source = driver.page_source if hasattr(driver, 'page_source') else ""
                    soup = BeautifulSoup(page_source, "html.parser")
                    block_check = is_page_blocked(soup, page_source)
                    if block_check["blocked"]:
                        logger.warning(f"[BLOCKED] Page {errored_page}: блокировка после таймаута → switching proxy")
                        cloudflare_block_count += 1
                except:
                    pass
                
                # Проверяем защиту от бесконечной смены прокси
                current_time = time.time()
                proxy_switch_times[:] = [t for t in proxy_switch_times if current_time - t < proxy_switch_period_seconds]
                if len(proxy_switch_times) >= max_proxy_switches_per_period:
                    logger.error(f"[ERROR] Превышен лимит смен прокси ({max_proxy_switches_per_period} за {proxy_switch_period_seconds} сек). Аварийное завершение.")
                    break
                
                try:
                    driver.quit()
                except:
                    pass
                logger.info(f"Ищем новый рабочий прокси для продолжения парсинга со страницы {errored_page}...")
                driver, start_from_index = get_driver_until_found(proxy_manager, start_from_index)
                proxy_switch_count += 1
                proxy_switch_times.append(time.time())
                logger.info(f"[OK] Новый прокси найден, продолжаем парсинг со страницы {errored_page}")
                continue
            except Exception as e:
                error_msg = str(e).lower()
                error_type = type(e).__name__
                
                # Проверяем, является ли ошибка связанной с прокси (прокси отказал в соединении)
                is_proxy_error = (
                    "proxyconnectfailure" in error_msg or
                    "proxy" in error_msg and ("refusing" in error_msg or "connection" in error_msg or "failed" in error_msg) or
                    "neterror" in error_msg and "proxy" in error_msg
                )
                
                logger.error(f"[ERROR] Ошибка при парсинге страницы {current_page}: {e}")
                if is_proxy_error:
                    logger.warning(f"[PROXY ERROR] Обнаружена ошибка прокси на странице {current_page}: {error_type}")
                
                # Пробуем перезагрузить страницу с тем же прокси (хотя бы один раз)
                logger.info(f"[RETRY] Повторная попытка загрузки страницы {current_page} с тем же прокси...")
                retry_success = False
                retry_is_proxy_error = False
                
                try:
                    soup, products_count = reload_page_if_needed(driver, page_url, max_retries=1)
                    products_retry = get_products_from_page_soup(soup)
                    if products_retry:
                        append_to_csv(TEMP_CSV_FILE, products_retry)
                        append_to_csv(TEMP_CSV_FILE, products_retry)
                        logger.info(f"[{thread_name}] Page {current_page} (retry): added {len(products_retry)} products")
                        total_collected += len(products_retry)
                        empty_pages_count = 0
                        pages_checked += 1
                        current_page += 1
                        time.sleep(random.uniform(2, 4))
                        retry_success = True
                except Exception as retry_error:
                    retry_error_msg = str(retry_error).lower()
                    retry_is_proxy_error = (
                        "proxyconnectfailure" in retry_error_msg or
                        "proxy" in retry_error_msg and ("refusing" in retry_error_msg or "connection" in retry_error_msg or "failed" in retry_error_msg) or
                        "neterror" in retry_error_msg and "proxy" in retry_error_msg
                    )
                    logger.warning(f"[WARNING] Повторная попытка страницы {current_page} не удалась: {retry_error}")
                    if retry_is_proxy_error:
                        logger.warning(f"[PROXY ERROR] При повторной попытке также ошибка прокси - меняем прокси")
                
                if retry_success:
                    continue
                
                # Если это ошибка прокси (и при повторной попытке тоже) - меняем прокси
                if is_proxy_error and retry_is_proxy_error:
                    logger.warning(f"[PROXY ERROR] Прокси отказал в соединении на странице {current_page} (и при повторной попытке тоже)")
                    logger.warning(f"[PROXY ERROR] Меняем прокси...")
                    errored_page = current_page
                    cloudflare_block_count += 1
                    
                    # Проверяем защиту от бесконечной смены прокси
                    current_time = time.time()
                    proxy_switch_times[:] = [t for t in proxy_switch_times if current_time - t < proxy_switch_period_seconds]
                    if len(proxy_switch_times) >= max_proxy_switches_per_period:
                        logger.error(f"[ERROR] Превышен лимит смен прокси ({max_proxy_switches_per_period} за {proxy_switch_period_seconds} сек). Аварийное завершение.")
                        break
                    
                    try:
                        driver.quit()
                    except:
                        pass
                    # Ищем новый прокси до тех пор пока не найдем
                    logger.info(f"Ищем новый рабочий прокси для продолжения парсинга со страницы {errored_page}...")
                    driver, start_from_index = get_driver_until_found(proxy_manager, start_from_index)
                    proxy_switch_count += 1
                    proxy_switch_times.append(time.time())
                    logger.info(f"[OK] Новый прокси найден, продолжаем парсинг со страницы {errored_page}")
                    continue
                
                # Запоминаем текущую страницу
                errored_page = current_page
                
                # Проверяем защиту от бесконечной смены прокси
                current_time = time.time()
                proxy_switch_times[:] = [t for t in proxy_switch_times if current_time - t < proxy_switch_period_seconds]
                if len(proxy_switch_times) >= max_proxy_switches_per_period:
                    logger.error(f"[ERROR] Превышен лимит смен прокси ({max_proxy_switches_per_period} за {proxy_switch_period_seconds} сек). Аварийное завершение.")
                    break
                
                try:
                    driver.quit()
                except:
                    pass
                # Ищем новый прокси до тех пор пока не найдем
                logger.info(f"Ищем новый рабочий прокси для продолжения парсинга со страницы {errored_page}...")
                driver, start_from_index = get_driver_until_found(proxy_manager, start_from_index)
                proxy_switch_count += 1
                proxy_switch_times.append(time.time())
                logger.info(f"[OK] Новый прокси найден, продолжаем парсинг со страницы {errored_page}")
                # Переходим к следующему циклу без изменения current_page
                continue
                
    finally:
        try:
            driver.quit()
        except:
            pass
    
    metrics = {
        "pages_checked": pages_checked,
        "proxy_switches": proxy_switch_count,
        "cloudflare_blocks": cloudflare_block_count,
        "max_empty_streak": max_empty_streak,
        "forced_proxy_updates": forced_proxy_updates,
        "page_count_attempts": page_count_attempts,
        "page_count_failures": page_count_failures,
        "fallback_mode": fallback_mode
    }
    logger.info(f"[{thread_name}] Итоговые метрики: {metrics}")
    
    return total_collected, metrics

def create_backup():
    """Создает бэкап основных файлов перед обновлением"""
    try:
        if os.path.exists(OUTPUT_FILE):
            shutil.copy2(OUTPUT_FILE, BACKUP_FILE)
            logger.info(f"Excel backup created: {BACKUP_FILE}")
        if os.path.exists(CSV_FILE):
            shutil.copy2(CSV_FILE, BACKUP_CSV)
            logger.info(f"CSV backup created: {BACKUP_CSV}")
    except Exception as e:
        logger.error(f"Error creating backup: {e}")

def convert_csv_to_excel(csv_path, excel_path):
    """Конвертирует CSV файл в Excel
    
    Args:
        csv_path: Путь к CSV файлу
        excel_path: Путь к выходному Excel файлу
    """
    try:
        if not os.path.exists(csv_path):
            logger.warning(f"[WARNING] CSV файл не найден: {csv_path}")
            return False
        
        logger.info(f"Конвертируем CSV в Excel: {csv_path} -> {excel_path}")
        
        # Создаем новый Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        
        # Читаем CSV и записываем в Excel
        with open(csv_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f, delimiter=';')
            for row in reader:
                ws.append(row)
        
        # Сохраняем Excel файл
        wb.save(excel_path)
        
        file_size = os.path.getsize(excel_path)
        logger.info(f"[OK] CSV конвертирован в Excel: {excel_path} (размер: {file_size} байт)")
        return True
        
    except Exception as e:
        logger.error(f"[ERROR] Ошибка при конвертации CSV в Excel: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return False

def finalize_output_files():
    """
    Финализирует временные файлы - перемещает CSV в основной и конвертирует в Excel.
    Это гарантирует, что старый файл не будет изменен при ошибке.
    """
    try:
        # Создаем бэкап старого файла перед заменой
        if os.path.exists(OUTPUT_FILE):
            create_backup()
        
        # Перемещаем временный CSV в основной
        if os.path.exists(TEMP_CSV_FILE):
            shutil.move(TEMP_CSV_FILE, CSV_FILE)
            logger.info(f"[OK] Временный CSV файл перемещен в основной: {CSV_FILE}")
            
            # Конвертируем CSV в Excel
            if convert_csv_to_excel(CSV_FILE, OUTPUT_FILE):
                logger.info(f"[OK] Excel файл создан из CSV: {OUTPUT_FILE}")
            else:
                logger.warning(f"[WARNING] Не удалось создать Excel файл из CSV")
        else:
            logger.warning("[WARNING]  Временный CSV файл не найден")
            
    except Exception as e:
        logger.error(f"[ERROR] Ошибка при финализации файлов: {e}")
        raise

def cleanup_temp_files():
    """Удаляет временные файлы в случае ошибки"""
    try:
        if os.path.exists(TEMP_CSV_FILE):
            os.remove(TEMP_CSV_FILE)
            logger.info(f"Временный CSV файл удален: {TEMP_CSV_FILE}")
        # Excel файл больше не создается во время парсинга, только в конце
        if os.path.exists(TEMP_OUTPUT_FILE):
            os.remove(TEMP_OUTPUT_FILE)
            logger.info(f"Временный Excel файл удален: {TEMP_OUTPUT_FILE}")
    except Exception as e:
        logger.warning(f"Не удалось удалить временные файлы: {e}")

def rename_log_file_by_status(status, total_products=0):
    """Переименовывает лог-файл с суффиксом на основе статуса
    
    Args:
        status: Статус парсинга ('done', 'insufficient_data', 'error')
        total_products: Количество собранных товаров (для определения успеха)
    """
    try:
        if not os.path.exists(LOG_FILE_PATH):
            logger.debug(f"Лог-файл не найден: {LOG_FILE_PATH}")
            return
        
        # Определяем суффикс на основе статуса
        if status == 'done' and total_products >= 100:
            suffix = "_success"
        elif status == 'insufficient_data':
            suffix = "_insufficient_data"
        elif status == 'error':
            suffix = "_failed"
        else:
            suffix = "_unknown"
        
        # Создаем новое имя файла
        base_name = os.path.splitext(LOG_FILE_PATH)[0]
        extension = os.path.splitext(LOG_FILE_PATH)[1]
        new_log_path = f"{base_name}{suffix}{extension}"
        
        # Переименовываем файл
        os.rename(LOG_FILE_PATH, new_log_path)
        logger.info(f" Лог-файл переименован: {os.path.basename(new_log_path)}")
        
    except Exception as e:
        logger.warning(f"[WARNING]  Не удалось переименовать лог-файл: {e}")

if __name__ == "__main__":
    script_name = "trast"
    logger.info("=== TRAST PARSER STARTED (PROXY-ONLY) ===")
    logger.info(f"Target URL: https://trast-zapchast.ru/shop/?_paged=1")
    logger.info(f"Start time: {datetime.now()}")
    TelegramNotifier.notify("[Trast] Update started")
    error_message = None
    
    try:
        start_time = datetime.now()
        set_script_start(script_name)
        logger.info("[OK] Database connection successful")
    except Exception as db_error:
        logger.warning(f"[WARNING]  Database connection failed: {db_error}, continuing without DB...")
        # Продолжаем без БД
        start_time = datetime.now()

    # Создаем временные файлы для записи (основной файл не трогаем)
    # Теперь создаем только CSV, Excel будет создан в конце из CSV
    try:
        create_new_csv(TEMP_CSV_FILE)
        logger.info("[OK] Создан временный CSV файл для записи данных")
    except Exception as file_error:
        logger.error(f"[ERROR] Ошибка при создании временных файлов: {file_error}")
        logger.error(f"[ERROR] Traceback: {traceback.format_exc()}")
        TelegramNotifier.notify(f"[Trast] Update failed — <code>{file_error}</code>")
        sys.exit(1)

    # Инициализируем прокси менеджер с фильтром по России
    logger.info("Step 1: Updating proxy list...")
    # Расширенный список стран: СНГ + Европа + Азия
    # СНГ: Россия, Беларусь, Казахстан, Армения, Азербайджан, Грузия, Кыргызстан, Молдова, Таджикистан, Туркменистан, Узбекистан, Украина
    CIS_COUNTRIES = ["RU", "BY", "KZ", "AM", "AZ", "GE", "KG", "MD", "TJ", "TM", "UZ", "UA"]
    # Европа (близко к России, часто работают с российскими сайтами)
    EUROPE_COUNTRIES = ["PL", "LT", "LV", "EE", "FI", "CZ", "SK", "HU", "RO", "BG", "DE", "NL", "SE", "FR"]
    # Азия (близко географически)
    ASIA_COUNTRIES = ["CN", "MN"]
    # Объединенный список
    EXTENDED_COUNTRIES = CIS_COUNTRIES + EUROPE_COUNTRIES + ASIA_COUNTRIES
    logger.info(f"Используем прокси из расширенного списка стран ({len(EXTENDED_COUNTRIES)} стран):")
    logger.info(f"  СНГ ({len(CIS_COUNTRIES)}): {', '.join(CIS_COUNTRIES)}")
    logger.info(f"  Европа ({len(EUROPE_COUNTRIES)}): {', '.join(EUROPE_COUNTRIES)}")
    logger.info(f"  Азия ({len(ASIA_COUNTRIES)}): {', '.join(ASIA_COUNTRIES)}")
    try:
        proxy_manager = ProxyManager(country_filter=EXTENDED_COUNTRIES)
        logger.info("[OK] ProxyManager инициализирован")
    except Exception as pm_error:
        logger.error(f"[ERROR] Ошибка при инициализации ProxyManager: {pm_error}")
        logger.error(f"[ERROR] Traceback: {traceback.format_exc()}")
        TelegramNotifier.notify(f"[Trast] Update failed — <code>{pm_error}</code>")
        sys.exit(1)
    
    # Стратегия ТОЛЬКО прокси - никакого прямого доступа
    logger.info("СТРАТЕГИЯ: ТОЛЬКО ПРОКСИ (расширенный список) - никакого прямого доступа!")
    logger.info("Прокси проверяются ТОЛЬКО на способность получить количество страниц с trast-zapchast.ru")
    
    # Принудительно обновляем список прокси при старте (чтобы загрузить с обоих источников)
    logger.info("Обновляем список прокси при старте (Proxifly + proxymania.su)...")
    try:
        if proxy_manager.download_proxies(force_update=True):
            logger.info("[OK] Список прокси успешно обновлен при старте")
        else:
            logger.warning("[WARNING] Не удалось обновить список прокси при старте, используем кэшированный")
    except Exception as update_error:
        logger.warning(f"[WARNING] Ошибка при обновлении прокси при старте: {update_error}")
        logger.warning("[WARNING] Продолжаем с кэшированным списком прокси")
    
    logger.info("============================================================")
    
    # Запускаем парсинг - он продолжается до 3 пустых страниц подряд
    # При блокировке/ошибке автоматически ищется новый прокси и парсинг продолжается
    producer_metrics = {}
    try:
        total_products, producer_metrics = producer(proxy_manager)
        logger.info(f"[OK] Producer завершился, собрано товаров: {total_products}")
    except Exception as producer_error:
        logger.error(f"[ERROR] Критическая ошибка в producer: {producer_error}")
        logger.error(f"[ERROR] Traceback: {traceback.format_exc()}")
        total_products = 0
        producer_metrics = {"failure": str(producer_error)}
        status = 'error'
        TelegramNotifier.notify(f"[Trast] Update failed — <code>{producer_error}</code>")
        cleanup_temp_files()
        try:
            set_script_end(script_name, status='error')
        except:
            pass
        # Переименовываем лог-файл перед выходом
        rename_log_file_by_status('error', total_products=0)
        sys.exit(1)

    # Определяем статус на основе результата
    if total_products == 0:
        status = 'insufficient_data'
        logger.critical(f"❗ Найдено 0 товаров")
    elif total_products >= 100:
        status = 'done'
    else:
        status = 'insufficient_data'
    
    try:
        if total_products >= 100:
            logger.info(f"[OK] Собрано {total_products} товаров - успешно!")
            # Перемещаем временные файлы в основные (старый файл сохраняется через бэкап)
            finalize_output_files()
            logger.info("[OK] Основные файлы обновлены успешно")
        else:
            logger.critical(f"❗ Недостаточно данных: {total_products} товаров")
            # Удаляем временные файлы - основной файл НЕ ТРОГАЕМ
            cleanup_temp_files()
            logger.info("[WARNING]  Временные файлы удалены, основной файл не изменен")
            # Восстанавливаем из бэкапа только если есть
            if os.path.exists(BACKUP_FILE) and not os.path.exists(OUTPUT_FILE):
                shutil.copy2(BACKUP_FILE, OUTPUT_FILE)
                logger.info("Excel восстановлен из бэкапа")
            if os.path.exists(BACKUP_CSV) and not os.path.exists(CSV_FILE):
                shutil.copy2(BACKUP_CSV, CSV_FILE)
                logger.info("CSV восстановлен из бэкапа")
    except Exception as e:
        logger.exception(f"[ERROR] Ошибка при финализации: {e}")
        status = 'error'
        error_message = str(e)
        # При ошибке удаляем временные файлы - основной файл остается нетронутым
        cleanup_temp_files()
        logger.info("[WARNING]  При ошибке временные файлы удалены, основной файл не изменен")

    duration = (datetime.now() - start_time).total_seconds()
    try:
        set_script_end(script_name, status=status)
    except Exception as db_end_error:
        logger.warning(f"[WARNING]  Ошибка при сохранении окончания в БД: {db_end_error}")
        if status != 'done':
            error_message = error_message or str(db_end_error)

    logger.info("============================================================")
    logger.info(f"Парсинг завершен! Всего собрано товаров: {total_products}")
    logger.info(f"Время выполнения: {round(duration, 2)} секунд")
    logger.info(f"Статус: {status}")
    logger.info(f"Товаров собрано: {total_products}")
    if producer_metrics:
        logger.info(f"Метрики парсинга: {producer_metrics}")
    logger.info("============================================================")

    metrics_suffix = ""
    if producer_metrics:
        pages_checked = producer_metrics.get("pages_checked")
        proxy_switches = producer_metrics.get("proxy_switches")
        cloudflare_blocks = producer_metrics.get("cloudflare_blocks")
        if pages_checked is not None or proxy_switches is not None or cloudflare_blocks is not None:
            parts = []
            if pages_checked is not None:
                parts.append(f"pages={pages_checked}")
            if proxy_switches is not None:
                parts.append(f"proxy_swaps={proxy_switches}")
            if cloudflare_blocks is not None:
                parts.append(f"cf_blocks={cloudflare_blocks}")
            metrics_suffix = " (" + ", ".join(parts) + ")"
    
    if status == 'done':
        TelegramNotifier.notify(f"[Trast] Update completed successfully — Duration: {duration:.2f}s, Products: {total_products}{metrics_suffix}")
    elif status == 'insufficient_data':
        TelegramNotifier.notify(f"[Trast] Update completed with insufficient data — Duration: {duration:.2f}s, Products: {total_products}{metrics_suffix}")
    else:
        failure_details = error_message or "Unknown error"
        TelegramNotifier.notify(f"[Trast] Update failed — <code>{failure_details}</code>")
    
    # Переименовываем лог-файл на основе статуса
    rename_log_file_by_status(status, total_products=total_products)

