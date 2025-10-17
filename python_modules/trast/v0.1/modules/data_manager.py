"""
Data management module for Trast parser.

Handles Excel/CSV operations, backups, and data validation.
"""

import os
import json
import shutil
import logging
import csv
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
from openpyxl import Workbook, load_workbook
from .config import TrastConfig

logger = logging.getLogger("trast.data_manager")


class DataWriter:
    """Handles writing data to Excel and CSV files."""
    
    def __init__(self):
        self.excel_file = TrastConfig.OUTPUT_FILE
        self.csv_file = TrastConfig.CSV_FILE
        self.total_products = 0
    
    def create_excel(self, path: Optional[str] = None):
        """Create new Excel file."""
        if path is None:
            path = self.excel_file
            
        if os.path.exists(path):
            os.remove(path)
            
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(["Manufacturer", "Article", "Description", "Price"])
        wb.save(path)
        logger.info(f"Created Excel file: {path}")
    
    def create_csv(self, path: Optional[str] = None):
        """Create new CSV file."""
        if path is None:
            path = self.csv_file
            
        if os.path.exists(path):
            os.remove(path)
            
        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f, delimiter=';')
            writer.writerow(["Manufacturer", "Article", "Description", "Price"])
        logger.info(f"Created CSV file: {path}")
    
    def append_to_excel(self, path: Optional[str] = None, product_list: List[Dict] = None):
        """Append products to Excel file."""
        if path is None:
            path = self.excel_file
        if product_list is None:
            product_list = []
            
        if not os.path.exists(path):
            self.create_excel(path)
            
        try:
            wb = load_workbook(path)
            ws = wb.active
            
            for product in product_list:
                ws.append([
                    product.get("manufacturer", ""),
                    product.get("article", ""),
                    product.get("description", ""),
                    product.get("price", {}).get("price", "")
                ])
            
            wb.save(path)
            self.total_products += len(product_list)
            
        except Exception as e:
            logger.error(f"Error writing to Excel: {e}")
        
        logger.info(f"Excel updated with {len(product_list)} records, file size: {os.path.getsize(path)} bytes")
    
    def append_to_csv(self, path: Optional[str] = None, product_list: List[Dict] = None):
        """Append products to CSV file."""
        if path is None:
            path = self.csv_file
        if product_list is None:
            product_list = []
            
        try:
            with open(path, 'a', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f, delimiter=';')
                for product in product_list:
                    writer.writerow([
                        product.get("manufacturer", ""),
                        product.get("article", ""),
                        product.get("description", ""),
                        product.get("price", {}).get("price", "")
                    ])
        except Exception as e:
            logger.error(f"Error writing to CSV: {e}")
    
    def get_stats(self) -> Dict[str, Any]:
        """Get data writing statistics."""
        return {
            'total_products': self.total_products,
            'excel_size': os.path.getsize(self.excel_file) if os.path.exists(self.excel_file) else 0,
            'csv_size': os.path.getsize(self.csv_file) if os.path.exists(self.csv_file) else 0
        }


class BackupManager:
    """Manages backup operations with metadata."""
    
    def __init__(self):
        self.backup_dir = TrastConfig.BACKUP_DIR
        self.excel_file = TrastConfig.OUTPUT_FILE
        self.csv_file = TrastConfig.CSV_FILE
        os.makedirs(self.backup_dir, exist_ok=True)
    
    def create_backup_with_metadata(self, excel_file: str, csv_file: str, 
                                   product_count: int, pages_processed: int) -> Dict[str, Any]:
        """Create backup with metadata."""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        backup_excel = os.path.join(self.backup_dir, f"trast_backup_{timestamp}.xlsx")
        backup_csv = os.path.join(self.backup_dir, f"trast_backup_{timestamp}.csv")
        
        # Copy files
        if os.path.exists(excel_file):
            shutil.copy2(excel_file, backup_excel)
        if os.path.exists(csv_file):
            shutil.copy2(csv_file, backup_csv)
        
        # Create metadata
        metadata = {
            'timestamp': timestamp,
            'product_count': product_count,
            'pages_processed': pages_processed,
            'excel_file': backup_excel,
            'csv_file': backup_csv,
            'excel_size': os.path.getsize(backup_excel) if os.path.exists(backup_excel) else 0,
            'csv_size': os.path.getsize(backup_csv) if os.path.exists(backup_csv) else 0
        }
        
        metadata_file = os.path.join(self.backup_dir, f"metadata_{timestamp}.json")
        with open(metadata_file, 'w', encoding='utf-8') as f:
            json.dump(metadata, f, indent=2, ensure_ascii=False)
        
        logger.info(f"Backup created in {self.backup_dir}: {product_count} products")
        return metadata
    
    def get_best_backup(self) -> Optional[Dict[str, Any]]:
        """Get the best backup based on product count."""
        import glob
        
        metadata_files = glob.glob(os.path.join(self.backup_dir, "metadata_*.json"))
        
        if not metadata_files:
            return None
        
        best_backup = None
        max_products = 0
        
        for meta_file in metadata_files:
            try:
                with open(meta_file, 'r', encoding='utf-8') as f:
                    metadata = json.load(f)
                    if metadata['product_count'] > max_products:
                        max_products = metadata['product_count']
                        best_backup = metadata
            except Exception as e:
                logger.error(f"Error reading metadata: {e}")
        
        return best_backup
    
    def incremental_backup(self, session_num: int, total_collected: int, pages_processed: int):
        """Create incremental backup after session."""
        logger.info(f"Incremental backup after session {session_num}")
        self.create_backup_with_metadata(
            self.excel_file, 
            self.csv_file, 
            total_collected, 
            pages_processed
        )
    
    def smart_restore(self, current_product_count: int, threshold_percent: float = 80) -> bool:
        """Smart restore if current result is below threshold."""
        best_backup = self.get_best_backup()
        
        if not best_backup:
            logger.info("No backups for comparison")
            return False
        
        best_count = best_backup['product_count']
        threshold = best_count * (threshold_percent / 100)
        
        logger.info(f"Comparison: Current={current_product_count}, Best={best_count}, Threshold={threshold:.0f}")
        
        if current_product_count < threshold:
            logger.warning(f"Current result below threshold, restoring...")
            
            if os.path.exists(best_backup['excel_file']):
                shutil.copy2(best_backup['excel_file'], self.excel_file)
                logger.info(f"Excel restored from backup")
            
            if os.path.exists(best_backup['csv_file']):
                shutil.copy2(best_backup['csv_file'], self.csv_file)
                logger.info(f"CSV restored from backup")
            
            return True
        else:
            logger.info("Current result is acceptable")
            return False
    
    def cleanup_old_backups(self, keep_count: int = 10):
        """Clean up old backups, keeping only the most recent ones."""
        import glob
        
        metadata_files = glob.glob(os.path.join(self.backup_dir, "metadata_*.json"))
        
        if len(metadata_files) <= keep_count:
            return
        
        # Sort by timestamp
        metadata_files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
        
        # Remove old backups
        for meta_file in metadata_files[keep_count:]:
            try:
                with open(meta_file, 'r', encoding='utf-8') as f:
                    metadata = json.load(f)
                
                # Remove associated files
                for file_key in ['excel_file', 'csv_file']:
                    file_path = metadata.get(file_key)
                    if file_path and os.path.exists(file_path):
                        os.remove(file_path)
                
                # Remove metadata file
                os.remove(meta_file)
                logger.debug(f"Removed old backup: {meta_file}")
                
            except Exception as e:
                logger.error(f"Error cleaning up backup {meta_file}: {e}")
    
    def get_backup_stats(self) -> Dict[str, Any]:
        """Get backup statistics."""
        import glob
        
        metadata_files = glob.glob(os.path.join(self.backup_dir, "metadata_*.json"))
        
        total_size = 0
        backup_count = len(metadata_files)
        
        for meta_file in metadata_files:
            try:
                with open(meta_file, 'r', encoding='utf-8') as f:
                    metadata = json.load(f)
                    total_size += metadata.get('excel_size', 0) + metadata.get('csv_size', 0)
            except Exception as e:
                logger.debug(f"Error reading backup metadata: {e}")
        
        return {
            'backup_count': backup_count,
            'total_size': total_size,
            'backup_dir': self.backup_dir
        }


class DataValidator:
    """Validates data quality and consistency."""
    
    @staticmethod
    def validate_product(product: Dict[str, Any]) -> bool:
        """Validate a single product."""
        required_fields = ['manufacturer', 'article', 'description', 'price']
        
        for field in required_fields:
            if field not in product:
                logger.warning(f"Product missing required field: {field}")
                return False
            
            if field == 'price':
                if not isinstance(product[field], dict) or 'price' not in product[field]:
                    logger.warning(f"Product has invalid price format: {product[field]}")
                    return False
            else:
                if not product[field] or not str(product[field]).strip():
                    logger.warning(f"Product has empty {field}")
                    return False
        
        return True
    
    @staticmethod
    def validate_product_list(products: List[Dict[str, Any]]) -> Tuple[List[Dict], List[Dict]]:
        """Validate product list and return valid/invalid products."""
        valid_products = []
        invalid_products = []
        
        for product in products:
            if DataValidator.validate_product(product):
                valid_products.append(product)
            else:
                invalid_products.append(product)
        
        logger.info(f"Validation: {len(valid_products)} valid, {len(invalid_products)} invalid products")
        return valid_products, invalid_products
    
    @staticmethod
    def get_data_quality_stats(products: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Get data quality statistics."""
        if not products:
            return {'total': 0, 'valid': 0, 'invalid': 0, 'quality_score': 0}
        
        valid_products, invalid_products = DataValidator.validate_product_list(products)
        
        quality_score = len(valid_products) / len(products) * 100
        
        return {
            'total': len(products),
            'valid': len(valid_products),
            'invalid': len(invalid_products),
            'quality_score': quality_score
        }
