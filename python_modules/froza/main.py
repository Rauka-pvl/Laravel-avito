import os
import csv
import sys
import shutil
import logging
import requests
from io import BytesIO
from datetime import datetime
from lxml import etree
from openpyxl import Workbook
import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
from bz_telebot.database_manager import set_script_start, set_script_end
from openpyxl.utils import get_column_letter

# === Import configuration ===
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "avito")))
from config import COMBINED_XML, LOG_DIR
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
from notification.main import TelegramNotifier

# === Froza credentials ===
LOGIN = "SIVF"
PASSWORD = "Jmb08OVg7b"

# === File paths for output and backup ===
OUTPUT_FILE = os.path.abspath(os.path.join(LOG_DIR, "..", "froza.xlsx"))
BACKUP_FILE = os.path.abspath(os.path.join(LOG_DIR, "..", "froza_backup.xlsx"))

# === Logging setup ===
logger = None
def setup_logging():
    global logger
    logger = logging.getLogger("froza")
    os.makedirs(LOG_DIR, exist_ok=True)
    log_subdir = os.path.abspath(os.path.join(LOG_DIR, "..", "logs-froza"))
    os.makedirs(log_subdir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = os.path.join(log_subdir, f"froza_{timestamp}.log")
    with open(log_path, "w", encoding="utf-8-sig") as f:
        f.write("")

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler(log_path, encoding="utf-8"),
            logging.StreamHandler(sys.stdout)
        ]
    )
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except AttributeError:
        pass
    return logging.getLogger("froza")

# === Parse XML response from Froza ===
def parse_xml_response(content):
    try:
        parser = etree.XMLParser(recover=True, encoding='utf-8')
        tree = etree.parse(BytesIO(content), parser)
        root = tree.getroot()
        return [
            {child.tag: child.text for child in elem}
            for elem in root
        ] if root is not None else []
    except Exception as e:
        logger.error(f"XML parsing error: {e}")
        return []

# === Fetch price list from Froza ===
def normalize_brand(value: str) -> str:
    return "".join(ch.lower() for ch in value if ch.isalnum())


def extract_item_brand(item: dict) -> str:
    brand_keys = ("make_name", "brand", "maker", "manufacturer", "brand_name")
    for key in brand_keys:
        value = item.get(key)
        if value:
            return value
    return ""


def get_price_list(code: str, brand: str = "") -> list:
    normalized_target = normalize_brand(brand) if brand else ""
    
    # Если указан бренд, ищем только по нему, без fallback на все бренды
    if brand:
        url = f"https://www.froza.ru/search_xml4.php?get=price_list&user={LOGIN}&password={PASSWORD}&code={code}&brand={brand}"
        try:
            logger.info(f"Requesting price list for: {code} (brand: {brand})")
            response = requests.get(url, timeout=5)
            response.raise_for_status()
            data = parse_xml_response(response.content)
            if not data:
                logger.warning(f"No data returned for code {code} with brand {brand}")
                return []

            # Фильтруем по нормализованному бренду
            filtered = [
                item for item in data
                if normalize_brand(extract_item_brand(item)) == normalized_target
            ]
            if filtered:
                logger.info(
                    f"Filtered {len(filtered)} offers matching brand '{brand}' out of {len(data)} total"
                )
                return filtered
            else:
                logger.warning(
                    f"No offers matched brand '{brand}' for code {code}. Returning empty list to avoid wrong prices."
                )
                # Не возвращаем все результаты, если указан бренд, но он не найден
                # Это предотвращает использование цен от других брендов
                return []
        except Exception as e:
            logger.error(f"Request error: {e}")
            return []
    
    # Если бренд не указан, возвращаем все результаты
    url = f"https://www.froza.ru/search_xml4.php?get=price_list&user={LOGIN}&password={PASSWORD}&code={code}"
    try:
        logger.info(f"Requesting price list for: {code} (any brand)")
        response = requests.get(url, timeout=5)
        response.raise_for_status()
        data = parse_xml_response(response.content)
        return data if data else []
    except Exception as e:
        logger.error(f"Request error: {e}")
        return []

# === Select the best offer (delivery <= 5 days) ===
def select_offer(data: list, oem: str = "", brand: str = "") -> tuple:
    fast, slow = [], []
    for item in data:
        try:
            delivery_time = int(item.get("delivery_time", 999))
            price = float(item.get("price", 9999999))
            if delivery_time <= 5:
                fast.append((price, item))
            else:
                slow.append((price, item))
        except Exception:
            continue

    if fast:
        fast.sort(key=lambda x: x[0])
        chosen = fast[0][1]
        logger.info(f"Fast offer: {chosen['price']} RUB, delivery {chosen['delivery_time']}-{chosen['delivery_time_guar']} days")
        return chosen, ""
    elif slow:
        slow.sort(key=lambda x: x[0])
        chosen = slow[0][1]
        logger.warning(f"Slow delivery: {chosen['price']} RUB, {chosen['delivery_time']}-{chosen['delivery_time_guar']} days")
        return chosen, "Delivery > 5 days"
    else:
        logger.warning(f"No offers found for {oem} ({brand})")
        return None, "No offers found"

# === Scan XML ad file ===
def scan_ads_file(filepath: str) -> list:
    tree = etree.parse(filepath)
    ads = tree.xpath("//Ad")
    total = len(ads)
    logger.info(f"Total ads found: {total}")

    results, processed, skipped = [], 0, 0
    for idx, ad in enumerate(ads, start=1):
        oem = ad.findtext("OEM")
        brand = ad.findtext("Brand")

        if oem and brand:
            logger.info(f"({idx}/{total}) Searching: OEM={oem}, Brand={brand}")
            prices = get_price_list(oem, brand)
            if prices:
                top_offers = prices[:5]
                for offer in top_offers:
                    results.append({
                        "Manufacturer": brand,
                        "Article": oem,
                        "Description": offer.get("description_rus", ""),
                        "Price": offer.get("price", ""),
                        "Delivery Time": f"{offer.get('delivery_time', '')}–{offer.get('delivery_time_guar', '')} days",
                        "Comment": (
                            "Delivery > 5 days"
                            if offer.get("delivery_time") and offer.get("delivery_time_guar")
                            and int(offer.get("delivery_time", "0")) > 5
                            else ""
                        )
                    })
                logger.info(f"Found {len(top_offers)} offers for OEM={oem}")
            else:
                logger.warning(f"No offers found for OEM={oem}, Brand={brand}")
                results.append({
                    "Manufacturer": brand,
                    "Article": oem,
                    "Description": "",
                    "Price": "",
                    "Delivery Time": "",
                    "Comment": "No offers found"
                })

            processed += 1
        else:
            logger.warning(f"Skipped ad ({idx}/{total}): missing OEM or Brand")
            skipped += 1

    logger.info(f"Processed: {processed}, Skipped: {skipped}")
    return results

# === Save data to XLSX ===
def save_to_xlsx(data: list, filename: str):
    if not data:
        logger.warning("No data to save.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Froza"
    headers = ["Manufacturer", "Article", "Description", "Price", "Delivery Time", "Comment"]
    ws.append(headers)

    for row in data:
        ws.append([row.get(h, "") for h in headers])

    for i, col in enumerate(ws.columns, start=1):
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[get_column_letter(i)].width = min(max_length + 2, 50)

    wb.save(filename)
    logger.info(f"Saved result to file: {filename}")

# === Create Excel backup ===
def create_backup():
    if os.path.exists(OUTPUT_FILE):
        shutil.copy2(OUTPUT_FILE, BACKUP_FILE)
        logger.info(f"Backup created: {BACKUP_FILE}")
    else:
        logger.info("Backup skipped: no output file found")

if __name__ == "__main__":
    logger = setup_logging()  # 🔧 инициализация логгера ДО try
    script_name = "froza"
    TelegramNotifier.notify("[Froza] Update started")
    start_time = datetime.now()
    set_script_start(script_name)

    try:
        create_backup()
        xlsx_filename = os.path.join(os.path.dirname(COMBINED_XML), "froza.xlsx")
        ads_data = scan_ads_file(COMBINED_XML)
        save_to_xlsx(ads_data, filename=xlsx_filename)

        end_time = datetime.now()
        duration = (end_time - start_time).total_seconds()
        set_script_end(script_name, status="done")

        TelegramNotifier.notify(
            f"[Froza] Update completed successfully — Duration: {duration:.2f}s"
        )

    except Exception as e:
        logging.exception("Error during processing:")
        set_script_end(script_name, status="failed")
        TelegramNotifier.notify(f"[Froza] Update failed — <code>{str(e)}</code>")
