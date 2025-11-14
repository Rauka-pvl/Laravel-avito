"""Excel writer for products"""

import sys
import threading
from pathlib import Path
from typing import List, Dict, Any
from openpyxl import Workbook, load_workbook

if str(Path(__file__).parent.parent) not in sys.path:
    sys.path.append(str(Path(__file__).parent.parent))

from logger import get_logger
from config import TEMP_EXCEL, EXCEL_FILE, TEMP_CSV, CSV_FILE
from storage.csv_writer import get_total_products

logger = get_logger("storage.excel_writer")

# Thread-safe lock for file operations
_file_lock = threading.Lock()


def create_excel_file(file_path: Path = None):
    """Create a new Excel file with headers"""
    if file_path is None:
        file_path = TEMP_EXCEL
    
    with _file_lock:
        if file_path.exists():
            file_path.unlink()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(["Manufacturer", "Article", "Description", "Price"])
        wb.save(file_path)
        
        logger.debug(f"Created Excel file: {file_path}")


def append_products(products: List[Dict[str, Any]], file_path: Path = None):
    """
    Append products to Excel file (thread-safe)
    
    Args:
        products: List of product dicts with keys: manufacturer, article, description, price
        file_path: Path to Excel file (defaults to TEMP_EXCEL)
    """
    if not products:
        return
    
    if file_path is None:
        file_path = TEMP_EXCEL
    
    with _file_lock:
        # Create file if it doesn't exist
        if not file_path.exists():
            create_excel_file(file_path)
        
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            for product in products:
                price_value = ""
                if isinstance(product.get("price"), dict):
                    price_value = product["price"].get("price", "")
                elif product.get("price"):
                    price_value = str(product["price"])
                
                ws.append([
                    product.get("manufacturer", ""),
                    product.get("article", ""),
                    product.get("description", ""),
                    price_value
                ])
            
            wb.save(file_path)
            logger.debug(f"Appended {len(products)} products to Excel")
        except Exception as e:
            logger.error(f"Error writing to Excel: {e}")
            raise


def convert_csv_to_excel(csv_path: Path = None, excel_path: Path = None):
    """
    Convert CSV file to Excel file
    
    Args:
        csv_path: Path to CSV file (defaults to TEMP_CSV)
        excel_path: Path to output Excel file (defaults to TEMP_EXCEL)
    """
    if csv_path is None:
        csv_path = TEMP_CSV
    if excel_path is None:
        excel_path = TEMP_EXCEL
    
    if not csv_path.exists():
        logger.warning(f"CSV file not found: {csv_path}")
        return
    
    import csv
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        
        with open(csv_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f, delimiter=';')
            for row in reader:
                ws.append(row)
        
        wb.save(excel_path)
        logger.info(f"Converted CSV to Excel: {csv_path} -> {excel_path}")
    except Exception as e:
        logger.error(f"Error converting CSV to Excel: {e}")
        raise


def finalize_output_files():
    """Move temporary files to final output files"""
    import shutil
    
    with _file_lock:
        try:
            # Backup existing files
            if EXCEL_FILE.exists():
                shutil.copy2(EXCEL_FILE, EXCEL_FILE.parent / f"{EXCEL_FILE.stem}_backup{EXCEL_FILE.suffix}")
            if CSV_FILE.exists():
                shutil.copy2(CSV_FILE, CSV_FILE.parent / f"{CSV_FILE.stem}_backup{CSV_FILE.suffix}")
            
            # Convert CSV to Excel if CSV exists
            if TEMP_CSV.exists():
                convert_csv_to_excel(TEMP_CSV, TEMP_EXCEL)
            
            # Move temp files to final
            if TEMP_EXCEL.exists():
                shutil.move(TEMP_EXCEL, EXCEL_FILE)
                logger.info(f"Excel file finalized: {EXCEL_FILE}")
            
            if TEMP_CSV.exists():
                shutil.move(TEMP_CSV, CSV_FILE)
                logger.info(f"CSV file finalized: {CSV_FILE}")
        except Exception as e:
            logger.error(f"Error finalizing output files: {e}")
            raise

